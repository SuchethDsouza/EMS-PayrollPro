from flask import (Flask, render_template, request, redirect, url_for, session, flash, jsonify)
import datetime, calendar, json, os, glob
from database import get_connection, setup_database, hash_password, get_db_path, setup_new_admin_db, DEMO_DB, BASE_DIR
from payroll_engine import calculate_all_payroll, get_payroll_summary, calculate_payroll
from email_service import send_email, log_email, get_email_logs

app = Flask(__name__)
app.secret_key = "payrollpro_secret_2024"

# ── Auth helpers ──────────────────────────────────────────────────────────────
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def hr_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("role") not in ("admin", "hr"):
            flash("Access denied. HR/Admin only.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            flash("Access denied. Admin only.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated


def find_user_in_all_dbs(identifier, password_hash):
    """
    Search all .db files in BASE_DIR for a matching user.
    Returns (user_row, db_username) or (None, None).
    identifier can be username or email.
    """
    import sqlite3
    # Collect all db files — demo first, then all payroll_*.db files
    db_files = [DEMO_DB] + sorted(glob.glob(os.path.join(BASE_DIR, "payroll_*.db")))
    # Remove demo from the glob results if it appears twice
    seen = set()
    unique_dbs = []
    for f in db_files:
        if f not in seen:
            seen.add(f)
            unique_dbs.append(f)

    for db_path in unique_dbs:
        if not os.path.exists(db_path):
            continue
        try:
            conn = sqlite3.connect(db_path, timeout=5)
            conn.row_factory = sqlite3.Row
            user = conn.execute(
                "SELECT * FROM users WHERE (username=? OR email=?) AND password_hash=?",
                (identifier, identifier, password_hash)
            ).fetchone()
            conn.close()
            if user:
                # Derive the session username from the db filename
                # payroll_sucheth.db → "sucheth", payroll_demo.db → "admin"
                basename = os.path.basename(db_path)
                if basename == "payroll_demo.db":
                    db_owner = user["username"]  # use actual username for demo
                else:
                    db_owner = basename.replace("payroll_", "").replace(".db", "")
                return dict(user), db_owner
        except Exception:
            continue
    return None, None

# ── Home (public landing page) ────────────────────────────────────────────────
@app.route("/")
def home():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return render_template("home.html")

# ── Auth ──────────────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    error = None
    if request.method == "POST":
        identifier = request.form.get("username", "").strip()
        password   = request.form.get("password", "")

        user, db_owner = find_user_in_all_dbs(identifier, hash_password(password))

        if user:
            if user.get("status") == "Inactive":
                error = "Your account has been deactivated. Please contact Admin."
            else:
                # Set username in session BEFORE anything else so
                # get_connection() automatically routes to the right DB
                session["username"]    = user["username"]
                session["user_id"]     = user["user_id"]
                session["role"]        = user["role"]
                session["employee_id"] = user["employee_id"]
                session["db_owner"]    = db_owner  # which DB file to use
                return redirect(url_for("dashboard"))
        else:
            error = "Invalid username / email or password."
    return render_template("login.html", error=error)


# ── Forgot Password ────────────────────────────────────────────────────────────
@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        identifier = request.form.get("identifier", "").strip()

        # Search across all DBs for the user
        import sqlite3 as _sq, glob as _glob
        found_user = None
        found_db   = None
        db_files   = [DEMO_DB] + sorted(_glob.glob(os.path.join(BASE_DIR, "payroll_*.db")))
        seen = set()
        for db_path in db_files:
            if db_path in seen or not os.path.exists(db_path):
                continue
            seen.add(db_path)
            try:
                c = _sq.connect(db_path, timeout=5)
                c.row_factory = _sq.Row
                u = c.execute("SELECT * FROM users WHERE username=? OR email=?",
                              (identifier, identifier)).fetchone()
                c.close()
                if u:
                    found_user = dict(u)
                    found_db   = db_path
                    break
            except Exception:
                continue

        if found_user:
            import secrets
            token   = secrets.token_urlsafe(32)
            expires = (datetime.datetime.now() + datetime.timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")
            c = _sq.connect(found_db, timeout=30)
            c.execute("UPDATE users SET reset_token=?, reset_expires=? WHERE user_id=?",
                      (token, expires, found_user["user_id"]))
            c.commit()
            c.close()

            reset_link      = url_for("reset_password", token=token, _external=True)
            recipient_email = found_user["email"]
            recipient_name  = found_user["username"].title()

            email_body = f"""You requested a password reset for your PayrollPro account.

Click the link below to reset your password. This link is valid for 1 hour.

{reset_link}

If you did not request this, please ignore this email.

Regards,
PayrollPro Security"""

            success, msg = send_email(
                recipient_email=recipient_email,
                recipient_name=recipient_name,
                subject="PayrollPro — Reset Your Password",
                body=email_body,
                sender_name="PayrollPro Security"
            )
            if success:
                flash(f"✅ Reset link sent to {recipient_email}. Check your inbox.", "success")
            else:
                flash(f"⚠️ Email failed ({msg}). Use this link: {reset_link}", "warning")
            return redirect(url_for("forgot_password"))
        else:
            flash("No account found with that username or email.", "danger")
    return render_template("forgot_password.html")


# ── Reset Password ─────────────────────────────────────────────────────────────
@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    import sqlite3 as _sq, glob as _glob
    found_user = None
    found_db   = None
    db_files   = [DEMO_DB] + sorted(_glob.glob(os.path.join(BASE_DIR, "payroll_*.db")))
    seen = set()
    for db_path in db_files:
        if db_path in seen or not os.path.exists(db_path):
            continue
        seen.add(db_path)
        try:
            c = _sq.connect(db_path, timeout=5)
            c.row_factory = _sq.Row
            u = c.execute("SELECT * FROM users WHERE reset_token=?", (token,)).fetchone()
            c.close()
            if u:
                found_user = dict(u)
                found_db   = db_path
                break
        except Exception:
            continue

    if not found_user:
        flash("Invalid or expired reset link.", "danger")
        return redirect(url_for("forgot_password"))

    if found_user["reset_expires"]:
        expires = datetime.datetime.strptime(found_user["reset_expires"], "%Y-%m-%d %H:%M:%S")
        if datetime.datetime.now() > expires:
            flash("Reset link has expired. Please request a new one.", "danger")
            return redirect(url_for("forgot_password"))

    error = None
    if request.method == "POST":
        new_pass = request.form.get("new_password", "")
        confirm  = request.form.get("confirm_password", "")
        if len(new_pass) < 6:
            error = "Password must be at least 6 characters."
        elif new_pass != confirm:
            error = "Passwords do not match."
        else:
            c = _sq.connect(found_db, timeout=30)
            c.execute("UPDATE users SET password_hash=?, reset_token=NULL, reset_expires=NULL WHERE user_id=?",
                      (hash_password(new_pass), found_user["user_id"]))
            c.commit()
            c.close()
            flash("✅ Password reset successfully! Please log in.", "success")
            return redirect(url_for("login"))

    return render_template("reset_password.html", token=token, error=error)


# ── Register ───────────────────────────────────────────────────────────────────
@app.route("/register", methods=["GET", "POST"])
def register():
    if "user_id" in session:
        return redirect(url_for("dashboard"))

    error = None

    if request.method == "POST":
        username  = request.form.get("username", "").strip().lower()
        full_name = request.form.get("full_name", "").strip()
        email     = request.form.get("email", "").strip().lower()
        password  = request.form.get("password", "")
        confirm   = request.form.get("confirm_password", "")

        # Block demo usernames
        if username in {"admin", "hr", "amit"}:
            error = "That username is reserved for demo accounts. Please choose another."
        elif not username or not email or not password or not full_name:
            error = "All fields are required."
        elif password != confirm:
            error = "Passwords do not match."
        elif len(password) < 6:
            error = "Password must be at least 6 characters."
        else:
            # Check across ALL existing DBs for username/email conflict
            existing_user, _ = find_user_in_all_dbs(username, hash_password(password))
            # Also check by email separately
            import glob as _glob
            email_taken = False
            for db_path in [DEMO_DB] + sorted(_glob.glob(os.path.join(BASE_DIR, "payroll_*.db"))):
                if not os.path.exists(db_path):
                    continue
                try:
                    import sqlite3 as _sq
                    _c = _sq.connect(db_path, timeout=5)
                    _c.row_factory = _sq.Row
                    row = _c.execute("SELECT user_id FROM users WHERE username=? OR email=?",
                                     (username, email)).fetchone()
                    _c.close()
                    if row:
                        email_taken = True
                        break
                except Exception:
                    pass

            if email_taken:
                error = "Username or email already taken."
            else:
                # Create a fresh DB for this new admin
                setup_new_admin_db(username)
                # Insert admin user into their new DB
                import sqlite3 as _sq
                db_path = get_db_path(username)
                conn = _sq.connect(db_path, timeout=30)
                conn.row_factory = _sq.Row
                conn.execute("PRAGMA journal_mode=WAL")
                conn.execute(
                    "INSERT INTO users (username, email, password_hash, role, status) VALUES (?,?,?,?,?)",
                    (username, email, hash_password(password), "admin", "Active")
                )
                conn.commit()
                conn.close()

                # ── Send welcome email with credentials to new admin ──────────
                login_url  = "http://127.0.0.1:5000/login"
                email_body = (
                    f"Welcome to PayrollPro!\n\n"
                    f"Your Admin account has been created successfully. "
                    f"Please keep these credentials confidential.\n\n"
                    f"{'─'*38}\n"
                    f"  Username  :  {username}\n"
                    f"  Password  :  {password}\n"
                    f"  Role      :  Admin (Company Owner)\n"
                    f"{'─'*38}\n\n"
                    f"Login at: {login_url}\n\n"
                    f"As Admin you can:\n"
                    f"• Add departments and HR executives\n"
                    f"• Manage employees, payroll, and attendance\n"
                    f"• Access all analytics and reports\n\n"
                    f"For security, please change your password after first login:\n"
                    f"My Profile → Change Password"
                )
                try:
                    send_email(
                        recipient_email=email,
                        recipient_name=full_name or username,
                        subject="Your PayrollPro Admin Account Credentials",
                        body=email_body,
                        sender_name="PayrollPro System"
                    )
                except Exception:
                    pass  # Don't block registration if email fails

                flash("✅ Admin account created! Check your email for login credentials.", "success")
                return redirect(url_for("login"))

    return render_template("register.html", error=error, blocked=False)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))

# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route("/dashboard")
@login_required
def dashboard():
    # Only redirect to employee dashboard if they have a linked employee record
    if session.get("role") == "employee" and session.get("employee_id"):
        return redirect(url_for("my_dashboard"))
    conn  = get_connection()
    today = str(datetime.date.today())
    now   = datetime.date.today()
    stats = {
        "total_emp":       conn.execute("SELECT COUNT(*) FROM employees WHERE status='Active'").fetchone()[0],
        "present":         conn.execute("SELECT COUNT(*) FROM attendance WHERE date=? AND status='Present'", (today,)).fetchone()[0],
        "absent":          conn.execute("SELECT COUNT(*) FROM attendance WHERE date=? AND status='Absent'",  (today,)).fetchone()[0],
        "on_leave":        conn.execute("SELECT COUNT(*) FROM attendance WHERE date=? AND status='Leave'",   (today,)).fetchone()[0],
        "pending_leaves":  conn.execute("SELECT COUNT(*) FROM leaves WHERE status='Pending'").fetchone()[0],
        "approved_leaves": conn.execute("SELECT COUNT(*) FROM leaves WHERE status='Approved'").fetchone()[0],
        "total_depts":     conn.execute("SELECT COUNT(*) FROM departments").fetchone()[0],
        "total_ann":       conn.execute("SELECT COUNT(*) FROM announcements WHERE is_active=1").fetchone()[0],
        "pending_payroll": conn.execute("""SELECT COUNT(*) FROM employees e
            LEFT JOIN payroll p ON p.employee_id=e.employee_id AND p.month=? AND p.year=?
            WHERE e.status='Active' AND p.payroll_id IS NULL""", (now.month,now.year)).fetchone()[0],
        "monthly_payroll": conn.execute(
            "SELECT COALESCE(SUM(net_salary),0) FROM payroll WHERE month=? AND year=?",
            (now.month, now.year)).fetchone()[0],
    }
    monthly_trend = []
    for i in range(5, -1, -1):
        m = ((now.month - 1 - i) % 12) + 1
        y = now.year if now.month - 1 - i >= 0 else now.year - 1
        avg = conn.execute("SELECT ROUND(AVG(overall_rating),1) FROM performance WHERE month=? AND year=?",
                           (m,y)).fetchone()[0] or 0
        monthly_trend.append({"month": calendar.month_abbr[m], "avg": avg})
    dept_perf = conn.execute("""SELECT d.dept_name, ROUND(AVG(p.overall_rating),1) as avg_rating
        FROM performance p JOIN employees e ON p.employee_id=e.employee_id
        JOIN departments d ON e.department_id=d.department_id WHERE p.year=?
        GROUP BY d.dept_name""", (now.year,)).fetchall()
    radar = conn.execute("""SELECT AVG(punctuality) as p, AVG(task_completion) as tc,
        AVG(teamwork) as tw, AVG(communication) as cm, AVG(initiative) as ini
        FROM performance WHERE year=?""", (now.year,)).fetchone()
    recent_att = conn.execute("""SELECT e.name, d.dept_name, a.status, a.check_in, a.check_out
        FROM attendance a JOIN employees e ON a.employee_id=e.employee_id
        JOIN departments d ON e.department_id=d.department_id
        WHERE a.date=? ORDER BY e.name LIMIT 8""", (today,)).fetchall()
    latest_anns = conn.execute("SELECT * FROM announcements WHERE is_active=1 ORDER BY ann_id DESC LIMIT 3").fetchall()
    conn.close()
    return render_template("dashboard.html", stats=stats,
                           trend=json.dumps(monthly_trend),
                           dept_perf=json.dumps([dict(r) for r in dept_perf]),
                           radar=json.dumps(dict(radar) if radar else {}),
                           recent_att=recent_att, latest_anns=latest_anns,
                           today=today, now=now)

# ── Employees ─────────────────────────────────────────────────────────────────
@app.route("/employees")
@hr_required
def employees():
    conn = get_connection()
    emps = conn.execute("""
        SELECT e.*, d.dept_name FROM employees e
        JOIN departments d ON e.department_id=d.department_id
        ORDER BY e.name
    """).fetchall()
    depts = conn.execute("SELECT * FROM departments").fetchall()
    conn.close()
    open_modal = request.args.get("open") == "add"
    return render_template("employees.html", employees=emps, departments=depts, open_modal=open_modal)

@app.route("/employees/add-form")
@hr_required
def add_employee_form():
    return redirect(url_for("employees") + "?open=add")

@app.route("/employees/add", methods=["POST"])
@hr_required
def add_employee():
    name     = request.form["name"]
    email    = request.form["email"]
    phone    = request.form["phone"]
    position = request.form["position"]
    dept_id  = request.form["department_id"]
    salary   = float(request.form["base_salary"])
    join     = request.form["join_date"]
    gender   = request.form.get("gender", "")
    # Role: if position contains "HR" → hr role, else employee
    role     = "hr" if "hr" in position.lower() or "human resource" in position.lower() else "employee"
    conn = None
    try:
        conn = get_connection()
        conn.execute(
            "INSERT INTO employees (name,email,phone,position,department_id,base_salary,join_date,gender,status) VALUES (?,?,?,?,?,?,?,?,'Active')",
            (name, email, phone, position, dept_id, salary, join, gender)
        )
        eid   = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        uname = name.split()[0].lower()    # First name as username
        pwd   = f"EMP{eid:03d}"            # e.g. EMP001, EMP002, EMP003

        conn.execute(
            "INSERT OR IGNORE INTO cl_balance (employee_id,year,total_cl,used_cl) VALUES (?,?,12,0)",
            (eid, datetime.date.today().year)
        )
        # Create login account
        conn.execute(
            "INSERT OR IGNORE INTO users (username,email,password_hash,role,employee_id,status) VALUES (?,?,?,?,?,'Active')",
            (uname, email, hash_password(pwd), role, eid)
        )
        conn.commit()

        # ── Credential email ──────────────────────────────────────────────────
        sender_role  = session.get("role", "admin")
        sender_label = "Admin" if sender_role == "admin" else "HR"
        login_url    = "http://127.0.0.1:5000/login"

        email_body = (
            f"Welcome to PayrollPro!\n\n"
            f"Your account has been created by {sender_label}. "
            f"Please keep these credentials confidential.\n\n"
            f"{'─'*38}\n"
            f"  Username  :  {uname}\n"
            f"  Password  :  {pwd}\n"
            f"{'─'*38}\n\n"
            f"Login at: {login_url}\n\n"
            f"For security, please change your password after first login:\n"
            f"My Profile → Change Password"
        )

        send_ok, msg = send_email(
            recipient_email=email,
            recipient_name=name,
            subject="Your PayrollPro Login Credentials",
            body=email_body,
            sender_name=f"PayrollPro {sender_label}"
        )

        if send_ok:
            flash(f"✅ '{name}' added! Credentials sent to {email}  (Username: {uname} / Password: {pwd})", "success")
        else:
            flash(f"✅ '{name}' added! Username: {uname}  Password: {pwd}  (Email failed: {msg})", "warning")

    except Exception as e:
        flash(f"Error adding employee: {str(e)}", "danger")
    finally:
        if conn:
            conn.close()
    return redirect(url_for("employees"))

@app.route("/employees/delete/<int:eid>")
@hr_required
def delete_employee(eid):
    conn = get_connection()
    conn.execute("UPDATE employees SET status='Inactive' WHERE employee_id=?", (eid,))
    conn.commit()
    conn.close()
    flash("Employee deactivated.", "info")
    return redirect(url_for("employees"))

# ── Attendance ────────────────────────────────────────────────────────────────
@app.route("/attendance")
@login_required
def attendance():
    date_str = request.args.get("date", str(datetime.date.today()))
    today    = datetime.date.today()
    conn     = get_connection()

    # Records for selected date
    records = conn.execute("""
        SELECT e.employee_id, e.name, d.dept_name,
               COALESCE(a.status,'—') as status,
               COALESCE(a.check_in,'—') as check_in,
               COALESCE(a.check_out,'—') as check_out
        FROM employees e
        JOIN departments d ON e.department_id=d.department_id
        LEFT JOIN attendance a ON a.employee_id=e.employee_id AND a.date=?
        WHERE e.status='Active' ORDER BY e.name
    """, (date_str,)).fetchall()

    # ── Stat card 1,2,3: TODAY's employee counts ─────────────────────────────
    today_str = str(today)
    present_today = conn.execute(
        "SELECT COUNT(DISTINCT employee_id) FROM attendance WHERE date=? AND status='Present'",
        (today_str,)).fetchone()[0]
    absent_today  = conn.execute(
        "SELECT COUNT(DISTINCT employee_id) FROM attendance WHERE date=? AND status='Absent'",
        (today_str,)).fetchone()[0]
    leave_today   = conn.execute(
        "SELECT COUNT(DISTINCT employee_id) FROM attendance WHERE date=? AND status='Leave'",
        (today_str,)).fetchone()[0]

    # ── Stat card 4: Monthly avg attendance % (weekdays only) ────────────────
    monthly = conn.execute("""
        SELECT status, COUNT(*) as cnt FROM attendance
        WHERE strftime('%m',date)=? AND strftime('%Y',date)=?
          AND CAST(strftime('%w', date) AS INTEGER) NOT IN (0, 6)
        GROUP BY status
    """, (f"{today.month:02d}", str(today.year))).fetchall()
    monthly_dict = {r["status"]: r["cnt"] for r in monthly}
    m_present = monthly_dict.get("Present", 0)
    m_absent  = monthly_dict.get("Absent",  0)
    m_leave   = monthly_dict.get("Leave",   0)
    m_total   = m_present + m_absent + m_leave
    att_pct   = round((m_present / m_total * 100), 1) if m_total else 0

    # ── Stat card 5: Weekday records this month ───────────────────────────────
    weekday_total = conn.execute("""
        SELECT COUNT(*) FROM attendance
        WHERE strftime('%m',date)=? AND strftime('%Y',date)=?
          AND CAST(strftime('%w', date) AS INTEGER) NOT IN (0, 6)
    """, (f"{today.month:02d}", str(today.year))).fetchone()[0]

    monthly_stats = {
        "present":       present_today,
        "absent":        absent_today,
        "on_leave":      leave_today,
        "att_pct":       att_pct,
        "total":         weekday_total,
    }

    # ── Per-employee attendance % — weekdays only ─────────────────────────────
    emp_stats = {}
    rows = conn.execute("""
        SELECT employee_id,
               SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present,
               COUNT(*) as total
        FROM attendance
        WHERE strftime('%m',date)=? AND strftime('%Y',date)=?
          AND CAST(strftime('%w', date) AS INTEGER) NOT IN (0, 6)
        GROUP BY employee_id
    """, (f"{today.month:02d}", str(today.year))).fetchall()
    for r in rows:
        pct = round((r["present"] / r["total"] * 100), 0) if r["total"] else 0
        emp_stats[r["employee_id"]] = {"present": r["present"], "total": r["total"], "pct": int(pct)}

    # ── Is selected date a weekend? ───────────────────────────────────────────
    try:
        selected_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        selected_date = today
    is_weekend  = selected_date.weekday() >= 5      # Sat=5, Sun=6
    is_future   = selected_date > today

    conn.close()
    return render_template("attendance.html", records=records, date_str=date_str,
                           monthly=json.dumps([dict(r) for r in monthly]),
                           monthly_stats=monthly_stats, emp_stats=emp_stats,
                           is_weekend=is_weekend, is_future=is_future)

def realistic_time(base_hour, variance_mins=20):
    """Generate a realistic check-in/out time with slight random variation."""
    import random as _r
    delta = _r.randint(-variance_mins, variance_mins)
    total = base_hour * 60 + delta
    h, m  = divmod(total, 60)
    return f"{h:02d}:{m:02d}"

@app.route("/attendance/mark", methods=["POST"])
@hr_required
def mark_attendance():
    emp_id   = request.form["employee_id"]
    date_str = request.form["date"]
    status   = request.form["status"]

    # ── Validation: block weekends and future dates ───────────────────────────
    try:
        mark_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        flash("Invalid date format.", "danger")
        return redirect(url_for("attendance"))
    if mark_date.weekday() >= 5:
        flash(f"❌ Cannot mark attendance on {mark_date.strftime('%A, %d %b %Y')} — it's a weekend.", "danger")
        return redirect(url_for("attendance", date=date_str))
    if mark_date > datetime.date.today():
        flash(f"❌ Cannot mark attendance for a future date ({mark_date.strftime('%d %b %Y')}).", "danger")
        return redirect(url_for("attendance", date=date_str))

    ci = realistic_time(9,  20) if status == "Present" else None
    co = realistic_time(18, 20) if status == "Present" else None
    conn = None
    try:
        conn = get_connection()
        ex = conn.execute("SELECT attendance_id FROM attendance WHERE employee_id=? AND date=?",
                          (emp_id, date_str)).fetchone()
        if ex:
            conn.execute("UPDATE attendance SET status=?,check_in=?,check_out=? WHERE attendance_id=?",
                         (status, ci, co, ex["attendance_id"]))
        else:
            conn.execute("INSERT INTO attendance (employee_id,date,status,check_in,check_out) VALUES (?,?,?,?,?)",
                         (emp_id, date_str, status, ci, co))
        conn.commit()
    finally:
        if conn: conn.close()
    dept    = request.args.get('dept', '')
    fstatus = request.args.get('status', '')
    return redirect(url_for("attendance", date=date_str, dept=dept, status=fstatus))


@app.route("/attendance/bulk-mark", methods=["POST"])
@hr_required
def attendance_bulk_mark():
    """Bulk mark: list of employee_ids OR 'all' flag."""
    date_str   = request.form.get("date", str(datetime.date.today()))
    status     = request.form.get("bulk_status", "Present")
    mark_all   = request.form.get("mark_all", "0") == "1"
    emp_ids    = request.form.getlist("emp_ids[]")

    # ── Validation: block weekends and future dates ───────────────────────────
    try:
        mark_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        flash("Invalid date format.", "danger")
        return redirect(url_for("attendance"))
    if mark_date.weekday() >= 5:
        flash(f"❌ Cannot mark attendance on {mark_date.strftime('%A, %d %b %Y')} — it's a weekend.", "danger")
        return redirect(url_for("attendance", date=date_str))
    if mark_date > datetime.date.today():
        flash(f"❌ Cannot bulk-mark attendance for a future date ({mark_date.strftime('%d %b %Y')}).", "danger")
        return redirect(url_for("attendance", date=date_str))

    conn = None
    try:
        conn = get_connection()
        if mark_all:
            rows = conn.execute("SELECT employee_id FROM employees WHERE status='Active'").fetchall()
            emp_ids = [str(r["employee_id"]) for r in rows]

        count = 0
        for eid in emp_ids:
            ci = realistic_time(9,  20) if status == "Present" else None
            co = realistic_time(18, 20) if status == "Present" else None
            ex = conn.execute("SELECT attendance_id FROM attendance WHERE employee_id=? AND date=?",
                              (eid, date_str)).fetchone()
            if ex:
                conn.execute("UPDATE attendance SET status=?,check_in=?,check_out=? WHERE attendance_id=?",
                             (status, ci, co, ex["attendance_id"]))
            else:
                conn.execute("INSERT INTO attendance (employee_id,date,status,check_in,check_out) VALUES (?,?,?,?,?)",
                             (eid, date_str, status, ci, co))
            count += 1
        conn.commit()
        flash(f"{count} employees marked as {status} for {date_str}.", "success")
    except Exception as e:
        flash(f"Error: {str(e)}", "danger")
    finally:
        if conn: conn.close()

    dept    = request.args.get('dept', '')
    fstatus = request.args.get('status', '')
    return redirect(url_for("attendance", date=date_str, dept=dept, status=fstatus))


@app.route("/attendance/export-excel")
@hr_required
def attendance_export_excel():
    """Export attendance for selected month/year to Excel."""
    import io, calendar as _cal
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file

    month = int(request.args.get("month", datetime.date.today().month))
    year  = int(request.args.get("year",  datetime.date.today().year))

    conn  = get_connection()
    rows  = conn.execute("""
        SELECT e.name, d.dept_name, e.position,
               SUM(CASE WHEN a.status='Present' THEN 1 ELSE 0 END) as present,
               SUM(CASE WHEN a.status='Absent'  THEN 1 ELSE 0 END) as absent,
               SUM(CASE WHEN a.status='Leave'   THEN 1 ELSE 0 END) as on_leave,
               COUNT(*) as total_marked
        FROM employees e
        JOIN departments d ON e.department_id=d.department_id
        LEFT JOIN attendance a ON a.employee_id=e.employee_id
            AND strftime('%m',a.date)=? AND strftime('%Y',a.date)=?
        WHERE e.status='Active'
        GROUP BY e.employee_id ORDER BY d.dept_name, e.name
    """, (f"{month:02d}", str(year))).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance {_cal.month_abbr[month]} {year}"
    ws.sheet_view.showGridLines = False

    purple  = PatternFill("solid", fgColor="2D2B55")
    hdr_fill= PatternFill("solid", fgColor="7F77DD")
    odd     = PatternFill("solid", fgColor="F2F3F8")
    even    = PatternFill("solid", fgColor="FFFFFF")
    green_f = PatternFill("solid", fgColor="1D9E75")
    thin    = Border(left=Side(style="thin",color="E2E4EC"),right=Side(style="thin",color="E2E4EC"),
                     top=Side(style="thin",color="E2E4EC"),bottom=Side(style="thin",color="E2E4EC"))

    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"PayrollPro — Attendance Report | {_cal.month_name[month]} {year}"
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = purple
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Headers
    headers = ["#","Employee","Department","Position","Present","Absent","On Leave","Attendance %"]
    widths  = [5, 22, 18, 18, 12, 12, 12, 15]
    for col,(h,w) in enumerate(zip(headers,widths),1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[2].height = 22

    # Data
    for i, r in enumerate(rows, 1):
        row_num = i + 2
        fill = odd if i % 2 else even
        pct  = round((r["present"] / r["total_marked"] * 100), 1) if r["total_marked"] else 0
        vals = [i, r["name"], r["dept_name"], r["position"],
                r["present"], r["absent"], r["on_leave"], f"{pct}%"]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if col not in (2,3,4) else "left", vertical="center")
            if col == 8:  # Attendance %
                cell.font = Font(color="155724" if pct >= 75 else "721C24", bold=True)
        ws.row_dimensions[row_num].height = 20

    # Totals
    tr = len(rows) + 3
    ws.merge_cells(f"A{tr}:D{tr}")
    ws[f"A{tr}"].value = "TOTAL"
    ws[f"A{tr}"].font  = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{tr}"].fill  = green_f
    ws[f"A{tr}"].alignment = Alignment(horizontal="center", vertical="center")
    for col, field in enumerate(["present","absent","on_leave"],5):
        cell = ws.cell(row=tr, column=col, value=sum(r[field] for r in rows))
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = green_f
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    ws.row_dimensions[tr].height = 26

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Attendance_{_cal.month_abbr[month]}_{year}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Payroll ───────────────────────────────────────────────────────────────────
@app.route("/payroll", methods=["GET", "POST"])
@hr_required
def payroll():
    today = datetime.date.today()
    month = int(request.form.get("month", today.month))
    year  = int(request.form.get("year", today.year))
    if request.method == "POST":
        calculate_all_payroll(month, year)
        flash(f"Payroll calculated for {calendar.month_name[month]} {year}!", "success")
    summary = get_payroll_summary(month, year)
    total   = sum(r["net_salary"] for r in summary)

    conn = get_connection()
    dept_data = conn.execute("""
        SELECT d.dept_name, SUM(p.net_salary) as total
        FROM payroll p
        JOIN employees e ON p.employee_id=e.employee_id
        JOIN departments d ON e.department_id=d.department_id
        WHERE p.month=? AND p.year=?
        GROUP BY d.dept_name
    """, (month, year)).fetchall()

    trend = []
    for m in range(1, 13):
        t = conn.execute("SELECT COALESCE(SUM(net_salary),0) FROM payroll WHERE month=? AND year=?",
                         (m, year)).fetchone()[0]
        trend.append({"month": calendar.month_abbr[m], "total": t})
    conn.close()

    months = list(calendar.month_name)[1:]
    years  = list(range(2020, today.year + 2))
    return render_template("payroll.html", summary=summary, total=total,
                           month=month, year=year, months=months, years=years,
                           dept_data=json.dumps([dict(r) for r in dept_data]),
                           trend=json.dumps(trend),
                           month_name=calendar.month_name[month])

# ── View Payroll (process payments) ──────────────────────────────────────────
@app.route("/view-payroll")
@hr_required
def view_payroll():
    today = datetime.date.today()
    month = int(request.args.get("month", today.month))
    year  = int(request.args.get("year",  today.year))
    conn  = get_connection()
    rows  = conn.execute("""
        SELECT p.payroll_id, p.employee_id, e.name, e.position,
               d.dept_name, p.basic_pay, p.deductions, p.net_salary,
               p.payment_status
        FROM payroll p
        JOIN employees e ON p.employee_id = e.employee_id
        JOIN departments d ON e.department_id = d.department_id
        WHERE p.month=? AND p.year=?
        ORDER BY d.dept_name, e.name
    """, (month, year)).fetchall()
    summary    = [dict(r) for r in rows]
    total      = sum(r["net_salary"]  for r in summary)
    paid_total = sum(r["net_salary"]  for r in summary if r["payment_status"] == "Paid")
    conn.close()
    return render_template("viewpayroll.html",
                           summary=summary, total=total, paid_total=paid_total,
                           month=month, year=year,
                           month_name=calendar.month_name[month],
                           now=today)

# ── Process single payment ────────────────────────────────────────────────────
@app.route("/process-payment", methods=["POST"])
@hr_required
def process_payment():
    employee_id = int(request.form.get("employee_id"))
    month       = int(request.form.get("month"))
    year        = int(request.form.get("year"))
    # ── Layer 3: Backend payment date lock ───────────────────────────────────
    today = datetime.date.today()
    is_past_month = (year, month) < (today.year, today.month)
    if not is_past_month and today.day < 28:
        days_left = 28 - today.day
        flash(f"⚠️ Salary payments are locked until the 28th of the month. "
              f"{days_left} day{'s' if days_left != 1 else ''} remaining.", "danger")
        return redirect(url_for("view_payroll", month=month, year=year))
    conn        = get_connection()
    conn.execute("""
        UPDATE payroll SET payment_status='Paid'
        WHERE employee_id=? AND month=? AND year=?
    """, (employee_id, month, year))
    conn.commit()
    name = conn.execute("SELECT name FROM employees WHERE employee_id=?", (employee_id,)).fetchone()["name"]
    conn.close()
    flash(f"✅ Payment released to {name} for {calendar.month_name[month]} {year}.", "success")
    return redirect(url_for("view_payroll", month=month, year=year))

# ── Process ALL pending payments ──────────────────────────────────────────────
@app.route("/process-all-payments", methods=["POST"])
@hr_required
def process_all_payments():
    month = int(request.form.get("month"))
    year  = int(request.form.get("year"))
    # ── Layer 3: Backend payment date lock ───────────────────────────────────
    today = datetime.date.today()
    is_past_month = (year, month) < (today.year, today.month)
    if not is_past_month and today.day < 28:
        days_left = 28 - today.day
        flash(f"⚠️ Bulk salary payment is locked until the 28th of the month. "
              f"{days_left} day{'s' if days_left != 1 else ''} remaining.", "danger")
        return redirect(url_for("view_payroll", month=month, year=year))
    conn  = get_connection()
    result = conn.execute("""
        UPDATE payroll SET payment_status='Paid'
        WHERE month=? AND year=? AND payment_status != 'Paid'
    """, (month, year))
    count = result.rowcount
    conn.commit()
    conn.close()
    flash(f"✅ {count} payment(s) released for {calendar.month_name[month]} {year}.", "success")
    return redirect(url_for("view_payroll", month=month, year=year))

# ── Leaves + CL ───────────────────────────────────────────────────────────────
@app.route("/leaves")
@login_required
def leaves():
    conn = get_connection()
    role = session.get("role")
    if role in ("admin", "hr"):
        year = datetime.date.today().year
        raw_leaves = conn.execute("""
            SELECT l.*, e.name as emp_name, e.email as emp_email,
                   d.dept_name,
                   COALESCE(cl.total_cl, 12)                          as cl_total,
                   COALESCE(cl.used_cl, 0)                            as cl_used,
                   COALESCE(cl.total_cl - cl.used_cl, 12)             as cl_remaining
            FROM leaves l
            JOIN employees e ON l.employee_id = e.employee_id
            LEFT JOIN departments d ON e.department_id = d.department_id
            LEFT JOIN cl_balance cl
                   ON cl.employee_id = l.employee_id AND cl.year = ?
            ORDER BY l.leave_id DESC
        """, (year,)).fetchall()
        all_leaves = [dict(r) for r in raw_leaves]
    else:
        eid = session.get("employee_id")
        all_leaves = conn.execute("""
            SELECT l.*, e.name as emp_name FROM leaves l
            JOIN employees e ON l.employee_id=e.employee_id
            WHERE l.employee_id=? ORDER BY l.leave_id DESC
        """, (eid,)).fetchall()

    if session.get("role") == "employee":
        emp_id = session.get("employee_id")
        cl_data = conn.execute("""
            SELECT e.name, d.dept_name, cl.total_cl, cl.used_cl, (cl.total_cl - cl.used_cl) as remaining
            FROM cl_balance cl
            JOIN employees e ON cl.employee_id = e.employee_id
            LEFT JOIN departments d ON e.department_id = d.department_id
            WHERE cl.year=? AND cl.employee_id=?
        """, (datetime.date.today().year, emp_id)).fetchall()
    else:
        cl_data = conn.execute("""
            SELECT e.name, d.dept_name, cl.total_cl, cl.used_cl, (cl.total_cl - cl.used_cl) as remaining
            FROM cl_balance cl
            JOIN employees e ON cl.employee_id = e.employee_id
            LEFT JOIN departments d ON e.department_id = d.department_id
            WHERE cl.year=?
            ORDER BY d.dept_name, e.name
        """, (datetime.date.today().year,)).fetchall()

    employees_list = conn.execute("SELECT employee_id, name FROM employees WHERE status='Active'").fetchall()
    conn.close()
    return render_template("leaves.html", leaves=all_leaves, cl_data=cl_data,
                           employees=employees_list)

@app.route("/leaves/apply", methods=["POST"])
@login_required
def apply_leave():
    if session.get("role") in ("admin", "hr"):
        emp_id = request.form["employee_id"]
    else:
        emp_id = session.get("employee_id")
    start      = request.form["start_date"]
    end        = request.form["end_date"]
    leave_type = request.form["leave_type"]
    reason     = request.form.get("reason", "")
    applied_on = str(datetime.date.today())

    if leave_type == "Casual":
        s = datetime.datetime.strptime(start, "%Y-%m-%d").date()
        e = datetime.datetime.strptime(end,   "%Y-%m-%d").date()
        days = sum(1 for d in range((e - s).days + 1)
                   if (s + datetime.timedelta(d)).weekday() < 5)
        conn = get_connection()
        cl = conn.execute("SELECT * FROM cl_balance WHERE employee_id=? AND year=?",
                          (emp_id, s.year)).fetchone()
        if cl and (cl["total_cl"] - cl["used_cl"]) < days:
            flash(f"Insufficient CL balance! Available: {cl['total_cl'] - cl['used_cl']} days.", "danger")
            conn.close()
            if session.get("role") == "employee":
                return redirect(url_for("leave_apply_form"))
            return redirect(url_for("leaves"))
        conn.close()

    conn = get_connection()
    conn.execute(
        "INSERT INTO leaves (employee_id,start_date,end_date,leave_type,reason,status,applied_on) VALUES (?,?,?,?,?,'Pending',?)",
        (emp_id, start, end, leave_type, reason, applied_on)
    )
    conn.commit()
    conn.close()
    flash("Leave application submitted!", "success")
    if session.get("role") == "employee":
        return redirect(url_for("leave_history"))
    return redirect(url_for("leaves"))

@app.route("/leaves/action/<int:lid>/<action>")
@hr_required
def leave_action(lid, action):
    status = "Approved" if action == "approve" else "Rejected"
    conn = get_connection()
    leave = conn.execute("SELECT * FROM leaves WHERE leave_id=?", (lid,)).fetchone()
    if leave and status == "Approved" and leave["leave_type"] == "Casual":
        s = datetime.datetime.strptime(leave["start_date"], "%Y-%m-%d").date()
        e = datetime.datetime.strptime(leave["end_date"],   "%Y-%m-%d").date()
        days = sum(1 for d in range((e - s).days + 1)
                   if (s + datetime.timedelta(d)).weekday() < 5)
        conn.execute("""
            UPDATE cl_balance SET used_cl = used_cl + ?
            WHERE employee_id=? AND year=?
        """, (days, leave["employee_id"], s.year))
    conn.execute("UPDATE leaves SET status=? WHERE leave_id=?", (status, lid))
    conn.commit()
    conn.close()
    flash(f"Leave {status.lower()} successfully.", "success")
    return redirect(url_for("leaves"))

# ── Employee: Leave History ────────────────────────────────────────────────────
@app.route("/leaves/history")
@login_required
def leave_history():
    if session.get("role") in ("admin", "hr"):
        return redirect(url_for("leaves"))
    emp_id = session.get("employee_id")
    year   = datetime.date.today().year
    conn   = get_connection()
    my_leaves = conn.execute("""
        SELECT l.*,
               COALESCE(cl.total_cl, 12)              AS cl_total,
               COALESCE(cl.used_cl,  0)               AS cl_used,
               COALESCE(cl.total_cl - cl.used_cl, 12) AS cl_remaining
        FROM leaves l
        LEFT JOIN cl_balance cl ON cl.employee_id = l.employee_id AND cl.year = ?
        WHERE l.employee_id = ?
        ORDER BY l.leave_id DESC
    """, (year, emp_id)).fetchall()
    conn.close()
    return render_template("leave_history.html", leaves=my_leaves)

# ── Employee: Apply for Leave page ────────────────────────────────────────────
@app.route("/leaves/apply-form")
@login_required
def leave_apply_form():
    if session.get("role") in ("admin", "hr"):
        return redirect(url_for("leaves"))
    emp_id = session.get("employee_id")
    year   = datetime.date.today().year
    conn   = get_connection()
    my_cl  = conn.execute("""
        SELECT cl.total_cl, cl.used_cl, (cl.total_cl - cl.used_cl) AS remaining
        FROM cl_balance cl WHERE cl.employee_id = ? AND cl.year = ?
    """, (emp_id, year)).fetchone()
    conn.close()
    return render_template("leave_apply.html", my_cl=my_cl)

# ── Performance ───────────────────────────────────────────────────────────────
@app.route("/performance")
@login_required
def performance():
    today = datetime.date.today()
    conn  = get_connection()

    # ── Accept month/year filter from GET params ──────────────────────────────
    try:
        month = int(request.args.get("month", today.month))
        year  = int(request.args.get("year",  today.year))
        dept_filter = request.args.get("dept", "")
        # Clamp to valid ranges
        month = max(1, min(12, month))
        year  = max(2020, min(today.year, year))
    except (ValueError, TypeError):
        month, year, dept_filter = today.month, today.year, ""

    # Base query for this month/year with optional dept filter
    dept_clause = "AND d.dept_name = :dept" if dept_filter else ""
    perf = conn.execute(f"""
        SELECT p.*, e.name as emp_name, d.dept_name
        FROM performance p
        JOIN employees e ON p.employee_id=e.employee_id
        JOIN departments d ON e.department_id=d.department_id
        WHERE p.month=:month AND p.year=:year {dept_clause}
        ORDER BY p.overall_rating DESC
    """, {"month": month, "year": year, "dept": dept_filter}).fetchall()

    radar_data = conn.execute("""
        SELECT AVG(punctuality) as p, AVG(task_completion) as tc,
               AVG(teamwork) as tw, AVG(communication) as cm, AVG(initiative) as ini
        FROM performance WHERE month=? AND year=?
    """, (month, year)).fetchone()

    monthly_trend = []
    for m in range(1, 13):
        avg = conn.execute(
            "SELECT ROUND(AVG(overall_rating),1) FROM performance WHERE month=? AND year=?",
            (m, year)
        ).fetchone()[0] or 0
        monthly_trend.append({"month": calendar.month_abbr[m], "avg": avg})

    dept_perf = conn.execute("""
        SELECT d.dept_name, ROUND(AVG(p.overall_rating),1) as avg_rating
        FROM performance p
        JOIN employees e ON p.employee_id=e.employee_id
        JOIN departments d ON e.department_id=d.department_id
        WHERE p.month=? AND p.year=?
        GROUP BY d.dept_name
    """, (month, year)).fetchall()

    top = conn.execute("""
        SELECT e.name, d.dept_name, ROUND(AVG(p.overall_rating),1) as avg_rating
        FROM performance p
        JOIN employees e ON p.employee_id=e.employee_id
        JOIN departments d ON e.department_id=d.department_id
        WHERE p.month=? AND p.year=?
        GROUP BY e.employee_id ORDER BY avg_rating DESC LIMIT 5
    """, (month, year)).fetchall()

    employees_list = conn.execute("SELECT employee_id, name FROM employees WHERE status='Active'").fetchall()
    total_emp = len(employees_list)

    # ── KPI: avg_rating & prev_avg ───────────────────────────────────────────
    avg_rating = round(
        conn.execute("SELECT AVG(overall_rating) FROM performance WHERE month=? AND year=?",
                     (month, year)).fetchone()[0] or 0, 1)
    prev_month = month - 1 if month > 1 else 12
    prev_year  = year if month > 1 else year - 1
    prev_avg   = round(
        conn.execute("SELECT AVG(overall_rating) FROM performance WHERE month=? AND year=?",
                     (prev_month, prev_year)).fetchone()[0] or 0, 1)

    # ── Per-employee averages ────────────────────────────────────────────────
    emp_avgs = conn.execute("""
        SELECT e.employee_id, e.name, d.dept_name,
               ROUND(AVG(p.overall_rating),1) as avg_r
        FROM performance p
        JOIN employees e ON p.employee_id=e.employee_id
        JOIN departments d ON e.department_id=d.department_id
        WHERE p.month=? AND p.year=?
        GROUP BY e.employee_id
    """, (month, year)).fetchall()

    high_perf_count = sum(1 for r in emp_avgs if (r["avg_r"] or 0) >= 80)
    at_risk_count   = sum(1 for r in emp_avgs if (r["avg_r"] or 0) < 65)
    high_perf_pct   = round(high_perf_count / total_emp * 100) if total_emp else 0

    # ── Reviews completed this month ─────────────────────────────────────────
    reviewed_this_month = conn.execute(
        "SELECT COUNT(DISTINCT employee_id) FROM performance WHERE month=? AND year=?",
        (month, year)).fetchone()[0] or 0
    review_pct = round(reviewed_this_month / total_emp * 100) if total_emp else 0

    # ── Bottom 5 performers ──────────────────────────────────────────────────
    bottom = conn.execute("""
        SELECT e.name, d.dept_name, ROUND(AVG(p.overall_rating),1) as avg_rating
        FROM performance p
        JOIN employees e ON p.employee_id=e.employee_id
        JOIN departments d ON e.department_id=d.department_id
        WHERE p.month=? AND p.year=?
        GROUP BY e.employee_id ORDER BY avg_rating ASC LIMIT 5
    """, (month, year)).fetchall()

    # ── Most improved employee ───────────────────────────────────────────────
    most_improved = None
    curr_ratings = conn.execute(
        "SELECT employee_id, overall_rating FROM performance WHERE month=? AND year=?",
        (month, year)).fetchall()
    prev_ratings = conn.execute(
        "SELECT employee_id, overall_rating FROM performance WHERE month=? AND year=?",
        (prev_month, prev_year)).fetchall()
    prev_map = {r["employee_id"]: r["overall_rating"] for r in prev_ratings}
    best_jump, best_eid = 0, None
    for r in curr_ratings:
        if r["employee_id"] in prev_map:
            jump = r["overall_rating"] - prev_map[r["employee_id"]]
            if jump > best_jump:
                best_jump, best_eid = jump, r["employee_id"]
    if best_eid:
        emp_row = conn.execute(
            "SELECT e.name, d.dept_name FROM employees e "
            "JOIN departments d ON e.department_id=d.department_id "
            "WHERE e.employee_id=?", (best_eid,)).fetchone()
        if emp_row:
            most_improved = {"name": emp_row["name"], "dept": emp_row["dept_name"],
                             "jump": round(best_jump, 1)}

    # ── Review coverage per department ───────────────────────────────────────
    dept_totals = conn.execute("""
        SELECT d.dept_name, COUNT(e.employee_id) as total
        FROM employees e JOIN departments d ON e.department_id=d.department_id
        WHERE e.status='Active' GROUP BY d.dept_name
    """).fetchall()
    dept_reviewed = conn.execute("""
        SELECT d.dept_name, COUNT(DISTINCT p.employee_id) as reviewed
        FROM performance p
        JOIN employees e ON p.employee_id=e.employee_id
        JOIN departments d ON e.department_id=d.department_id
        WHERE p.month=? AND p.year=? GROUP BY d.dept_name
    """, (month, year)).fetchall()
    reviewed_map = {r["dept_name"]: r["reviewed"] for r in dept_reviewed}
    review_coverage = [{"dept": r["dept_name"],
                        "reviewed": reviewed_map.get(r["dept_name"], 0),
                        "total": r["total"]} for r in dept_totals]

    # ── Scatter: attendance % vs performance ─────────────────────────────────
    scatter = []
    for r in emp_avgs:
        total_days = conn.execute(
            "SELECT COUNT(*) FROM attendance WHERE employee_id=? AND strftime('%Y',date)=? AND strftime('%m',date)=?",
            (r["employee_id"], str(year), f"{month:02d}")).fetchone()[0] or 0
        present_days = conn.execute(
            "SELECT COUNT(*) FROM attendance WHERE employee_id=? AND status='Present' AND strftime('%Y',date)=? AND strftime('%m',date)=?",
            (r["employee_id"], str(year), f"{month:02d}")).fetchone()[0] or 0
        att_pct = round(present_days / total_days * 100, 1) if total_days else 0
        scatter.append({"name": r["name"], "dept": r["dept_name"],
                        "x": att_pct, "y": r["avg_r"] or 0})

    # ── Smart insights ───────────────────────────────────────────────────────
    insights = []
    for dp in dept_perf:
        d = dict(dp)
        if d["avg_rating"] and d["avg_rating"] >= 85:
            insights.append({"icon": "🏆",
                "text": "{} is the top performing dept with avg {}".format(d["dept_name"], d["avg_rating"])})
        elif d["avg_rating"] and d["avg_rating"] < 70:
            insights.append({"icon": "⚠️",
                "text": "{} is below target at {} — consider a review".format(d["dept_name"], d["avg_rating"])})
    if prev_avg > 0:
        if avg_rating > prev_avg:
            insights.append({"icon": "📈",
                "text": "Company rating improved by {} pts vs last month".format(round(avg_rating - prev_avg, 1))})
        elif avg_rating < prev_avg:
            insights.append({"icon": "📉",
                "text": "Company rating dropped by {} pts vs last month".format(round(prev_avg - avg_rating, 1))})
    if review_pct < 50:
        insights.append({"icon": "📋",
            "text": "Only {}/{} employees reviewed this month — action needed".format(reviewed_this_month, total_emp)})
    elif review_pct >= 90:
        insights.append({"icon": "✅",
            "text": "{}% of employees reviewed this month — excellent coverage".format(review_pct)})
    if most_improved:
        insights.append({"icon": "⬆️",
            "text": "{} showed the biggest improvement (+{} pts)".format(
                most_improved["name"], most_improved["jump"])})

    # ── Departments list for filter dropdown ────────────────────────────────
    depts = conn.execute("SELECT dept_name FROM departments ORDER BY dept_name").fetchall()

    # ── Salary cost by department (for selected month/year) ─────────────────
    salary_by_dept = conn.execute("""
        SELECT d.dept_name, ROUND(SUM(p.net_salary), 2) as total_salary
        FROM payroll p
        JOIN employees e ON p.employee_id = e.employee_id
        JOIN departments d ON e.department_id = d.department_id
        WHERE p.month=? AND p.year=?
        GROUP BY d.dept_name ORDER BY total_salary DESC
    """, (month, year)).fetchall()

    # ── Monthly payroll trend for selected year ──────────────────────────────
    payroll_trend = []
    for m in range(1, 13):
        total = conn.execute(
            "SELECT ROUND(SUM(net_salary),2) FROM payroll WHERE month=? AND year=?",
            (m, year)
        ).fetchone()[0] or 0
        payroll_trend.append({"month": calendar.month_abbr[m], "total": total})

    # ── Attendance pie data for selected month/year ──────────────────────────
    att_present = conn.execute(
        "SELECT COUNT(*) FROM attendance WHERE status='Present' AND strftime('%m',date)=? AND strftime('%Y',date)=?",
        (f"{month:02d}", str(year))).fetchone()[0] or 0
    att_absent = conn.execute(
        "SELECT COUNT(*) FROM attendance WHERE status='Absent' AND strftime('%m',date)=? AND strftime('%Y',date)=?",
        (f"{month:02d}", str(year))).fetchone()[0] or 0
    att_leave = conn.execute(
        "SELECT COUNT(*) FROM attendance WHERE status='Leave' AND strftime('%m',date)=? AND strftime('%Y',date)=?",
        (f"{month:02d}", str(year))).fetchone()[0] or 0

    conn.close()
    radar = dict(radar_data) if radar_data else {}
    return render_template("performance.html",
        perf=perf, radar=json.dumps(radar),
        monthly_trend=json.dumps(monthly_trend),
        dept_perf=json.dumps([dict(r) for r in dept_perf]),
        top=top, employees=employees_list, now=today,
        avg_rating=avg_rating, prev_avg=prev_avg,
        high_perf_count=high_perf_count, high_perf_pct=high_perf_pct,
        at_risk_count=at_risk_count,
        reviewed_this_month=reviewed_this_month, review_pct=review_pct,
        total_emp=total_emp,
        bottom=bottom, most_improved=most_improved,
        review_coverage=review_coverage,
        scatter=json.dumps(scatter),
        insights=insights,
        depts=depts,
        salary_by_dept=json.dumps([dict(r) for r in salary_by_dept]),
        payroll_trend=json.dumps(payroll_trend),
        att_pie=json.dumps({"present": att_present, "absent": att_absent, "leave": att_leave}),
        month=month, year=year, dept_filter=dept_filter)

@app.route("/performance/add", methods=["POST"])
@hr_required
def add_performance():
    emp_id  = request.form["employee_id"]
    month   = int(request.form["month"])
    year    = int(request.form["year"])
    p   = int(request.form["punctuality"])
    tc  = int(request.form["task_completion"])
    tw  = int(request.form["teamwork"])
    cm  = int(request.form["communication"])
    ini = int(request.form["initiative"])
    overall  = round((p + tc + tw + cm + ini) / 5, 1)
    remarks  = request.form.get("remarks", "")
    reviewer = session.get("username", "HR")
    conn = get_connection()
    conn.execute("""
        INSERT OR REPLACE INTO performance
        (employee_id,month,year,punctuality,task_completion,teamwork,communication,initiative,overall_rating,remarks,reviewed_by)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (emp_id, month, year, p, tc, tw, cm, ini, overall, remarks, reviewer))
    conn.commit()
    conn.close()
    flash("Performance review saved!", "success")
    return redirect(url_for("performance"))

@app.route("/performance/employee/<int:eid>")
@login_required
def employee_performance(eid):
    conn = get_connection()
    emp = conn.execute("SELECT * FROM employees WHERE employee_id=?", (eid,)).fetchone()
    records = conn.execute("""
        SELECT * FROM performance WHERE employee_id=? ORDER BY year DESC, month DESC
    """, (eid,)).fetchall()
    trend = [{"month": calendar.month_abbr[r["month"]], "rating": r["overall_rating"]} for r in records]
    trend.reverse()
    conn.close()
    return render_template("employee_performance.html", emp=emp, records=records,
                           trend=json.dumps(trend))

# ── Email ─────────────────────────────────────────────────────────────────────
@app.route("/email", methods=["GET", "POST"])
@hr_required
def email_page():
    conn = get_connection()
    employees_list = conn.execute("SELECT employee_id, name, email FROM employees WHERE status='Active'").fetchall()
    conn.close()
    logs = get_email_logs()
    sent = False
    if request.method == "POST":
        emp_id  = int(request.form["recipient_id"])
        subject = request.form["subject"]
        body    = request.form["body"]
        conn = get_connection()
        emp = conn.execute("SELECT name, email FROM employees WHERE employee_id=?", (emp_id,)).fetchone()
        conn.close()
        success, msg = send_email(emp["email"], emp["name"], subject, body,
                                  sender_name=session.get("username", "HR"))
        status = "Sent" if success else "Failed"
        log_email(session.get("username"), emp_id, subject, body, status)
        flash(msg, "success" if success else "warning")
        logs = get_email_logs()
        sent = True
    return render_template("email.html", employees=employees_list, logs=logs, sent=sent)

@app.route("/api/employee/<int:eid>/performance")
@login_required
def api_emp_perf(eid):
    conn = get_connection()
    rows = conn.execute("""
        SELECT month, year, punctuality, task_completion, teamwork, communication, initiative, overall_rating
        FROM performance WHERE employee_id=? ORDER BY year, month
    """, (eid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ── Reports & Exports ─────────────────────────────────────────────────────────
@app.route("/reports")
@hr_required
def reports():
    today  = datetime.date.today()
    months = list(calendar.month_name)[1:]
    years  = list(range(2020, today.year + 2))
    conn   = get_connection()
    employees_list = conn.execute(
        "SELECT employee_id, name FROM employees WHERE status='Active' ORDER BY name"
    ).fetchall()
    conn.close()
    return render_template("reports.html", months=months, years=years,
                           employees=employees_list)

@app.route("/reports/payslip/<int:emp_id>/<int:month>/<int:year>")
@login_required
def download_payslip(emp_id, month, year):
    if session.get("role") == "employee" and session.get("employee_id") != emp_id:
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))
    from export_engine import generate_payslip_pdf
    from flask import send_file
    buf = generate_payslip_pdf(emp_id, month, year)
    if not buf:
        flash("No payroll data. Calculate payroll first.", "warning")
        return redirect(url_for("reports"))
    conn = get_connection()
    emp  = conn.execute("SELECT name FROM employees WHERE employee_id=?", (emp_id,)).fetchone()
    conn.close()
    name = emp["name"].replace(" ", "_") if emp else str(emp_id)
    return send_file(buf, as_attachment=True,
                     download_name=f"Payslip_{name}_{calendar.month_abbr[month]}_{year}.pdf",
                     mimetype="application/pdf")

@app.route("/reports/excel/<int:month>/<int:year>")
@hr_required
def download_excel(month, year):
    from export_engine import generate_payroll_excel
    from flask import send_file
    buf = generate_payroll_excel(month, year)
    return send_file(buf, as_attachment=True,
                     download_name=f"Payroll_{calendar.month_abbr[month]}_{year}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Profile ───────────────────────────────────────────────────────────────────
@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    emp_id = session.get("employee_id")
    conn   = get_connection()
    emp = att_summary = cl_info = perf_latest = None
    admin_stats = None

    # ── Admin profile: fetch org-wide stats ───────────────────────────────────
    if session.get("role") == "admin" and not emp_id:
        today = str(datetime.date.today())
        now   = datetime.date.today()

        # Handle password change POST for admin
        if request.method == "POST" and "current_password" in request.form:
            pass  # handled by change_password route

        # Handle email/username update POST for admin
        if request.method == "POST" and "admin_email" in request.form:
            new_email = request.form.get("admin_email", "").strip()
            conn.execute("UPDATE users SET email=? WHERE user_id=?",
                         (new_email, session["user_id"]))
            conn.commit()
            flash("Profile updated!", "success")

        # Fetch the admin user record for email display
        admin_user = conn.execute(
            "SELECT * FROM users WHERE user_id=?", (session["user_id"],)
        ).fetchone()

        admin_stats = {
            "total_emp":       conn.execute("SELECT COUNT(*) FROM employees WHERE status='Active'").fetchone()[0],
            "total_depts":     conn.execute("SELECT COUNT(*) FROM departments").fetchone()[0],
            "present_today":   conn.execute("SELECT COUNT(*) FROM attendance WHERE date=? AND status='Present'", (today,)).fetchone()[0],
            "absent_today":    conn.execute("SELECT COUNT(*) FROM attendance WHERE date=? AND status='Absent'",  (today,)).fetchone()[0],
            "pending_leaves":  conn.execute("SELECT COUNT(*) FROM leaves WHERE status='Pending'").fetchone()[0],
            "approved_leaves": conn.execute("SELECT COUNT(*) FROM leaves WHERE status='Approved'").fetchone()[0],
            "total_ann":       conn.execute("SELECT COUNT(*) FROM announcements WHERE is_active=1").fetchone()[0],
            "monthly_payroll": conn.execute(
                "SELECT COALESCE(SUM(net_salary),0) FROM payroll WHERE month=? AND year=?",
                (now.month, now.year)).fetchone()[0],
            "pending_payroll": conn.execute("""SELECT COUNT(*) FROM employees e
                LEFT JOIN payroll p ON p.employee_id=e.employee_id AND p.month=? AND p.year=?
                WHERE e.status='Active' AND p.payroll_id IS NULL""",
                (now.month, now.year)).fetchone()[0],
            "avg_performance": conn.execute(
                "SELECT ROUND(AVG(overall_rating),1) FROM performance WHERE year=?",
                (now.year,)).fetchone()[0] or 0,
            "total_users":     conn.execute("SELECT COUNT(*) FROM users WHERE status='Active'").fetchone()[0],
            "month_name":      calendar.month_name[now.month],
            "year":            now.year,
            "admin_email":     admin_user["email"] if admin_user else "",
        }

        conn.close()
        return render_template("profile.html", emp=None, att=None,
                               cl=None, perf=None, admin_stats=admin_stats)

    # ── HR / Employee profile ─────────────────────────────────────────────────
    if emp_id:
        emp = conn.execute("""
            SELECT e.*, d.dept_name FROM employees e
            JOIN departments d ON e.department_id=d.department_id
            WHERE e.employee_id=?
        """, (emp_id,)).fetchone()

        if request.method == "POST":
            conn.execute("UPDATE employees SET phone=?, email=? WHERE employee_id=?",
                         (request.form.get("phone",""), request.form.get("email",""), emp_id))
            conn.commit()
            flash("Profile updated!", "success")
            emp = conn.execute("""
                SELECT e.*, d.dept_name FROM employees e
                JOIN departments d ON e.department_id=d.department_id
                WHERE e.employee_id=?
            """, (emp_id,)).fetchone()

        today = datetime.date.today()
        att_summary = conn.execute("""
            SELECT SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present,
                   SUM(CASE WHEN status='Absent'  THEN 1 ELSE 0 END) as absent,
                   SUM(CASE WHEN status='Leave'   THEN 1 ELSE 0 END) as on_leave,
                   COUNT(*) as total
            FROM attendance WHERE employee_id=?
            AND strftime('%m',date)=? AND strftime('%Y',date)=?
        """, (emp_id, f"{today.month:02d}", str(today.year))).fetchone()

        cl_info = conn.execute(
            "SELECT * FROM cl_balance WHERE employee_id=? AND year=?",
            (emp_id, today.year)
        ).fetchone()

        perf_latest = conn.execute("""
            SELECT * FROM performance WHERE employee_id=?
            ORDER BY year DESC, month DESC LIMIT 1
        """, (emp_id,)).fetchone()

    conn.close()
    return render_template("profile.html", emp=emp, att=att_summary,
                           cl=cl_info, perf=perf_latest, admin_stats=None)

# ── Change Password ────────────────────────────────────────────────────────────
@app.route("/change-password", methods=["POST"])
@login_required
def change_password():
    current  = request.form.get("current_password", "")
    new_pass = request.form.get("new_password", "")
    confirm  = request.form.get("confirm_password", "")
    conn = get_connection()
    user = conn.execute("SELECT * FROM users WHERE user_id=? AND password_hash=?",
                        (session["user_id"], hash_password(current))).fetchone()
    if not user:
        flash("Current password is incorrect.", "danger")
    elif new_pass != confirm:
        flash("New passwords do not match.", "danger")
    elif len(new_pass) < 6:
        flash("Password must be at least 6 characters.", "danger")
    else:
        conn.execute("UPDATE users SET password_hash=? WHERE user_id=?",
                     (hash_password(new_pass), session["user_id"]))
        conn.commit()
        flash("Password changed successfully!", "success")
    conn.close()
    return redirect(url_for("profile"))

# ── Context processor ─────────────────────────────────────────────────────────
@app.context_processor
def inject_globals():
    return {"now": datetime.date.today(), "calendar": calendar}

# ── Entry point ───────────────────────────────────────────────────────────────

# ═══════════════════════════════════════════════════════════════════
# NEW ROUTES — Restructured layout
# ═══════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════
# NEW ROUTES — Restructured layout
# ═══════════════════════════════════════════════════════════════════

# ── Employee Dashboard (user role) ─────────────────────────────────
@app.route("/my-dashboard")
@login_required
def my_dashboard():
    emp_id = session.get("employee_id")
    # Don't redirect to dashboard — that causes an infinite loop for accounts without employee records
    if not emp_id:
        return render_template("my_dashboard.html", emp=None, att=None, cl=None,
                               perf=None, payroll=None, anns=[])
    conn = get_connection()
    emp = conn.execute("""
        SELECT e.*, d.dept_name FROM employees e
        JOIN departments d ON e.department_id=d.department_id
        WHERE e.employee_id=?
    """, (emp_id,)).fetchone()
    today = datetime.date.today()
    att = conn.execute("""
        SELECT SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present,
               SUM(CASE WHEN status='Absent'  THEN 1 ELSE 0 END) as absent,
               SUM(CASE WHEN status='Leave'   THEN 1 ELSE 0 END) as on_leave
        FROM attendance WHERE employee_id=?
        AND strftime('%m',date)=? AND strftime('%Y',date)=?
    """, (emp_id, f"{today.month:02d}", str(today.year))).fetchone()
    cl = conn.execute("SELECT * FROM cl_balance WHERE employee_id=? AND year=?",
                      (emp_id, today.year)).fetchone()
    perf = conn.execute("""SELECT * FROM performance WHERE employee_id=?
        ORDER BY year DESC, month DESC LIMIT 1""", (emp_id,)).fetchone()
    payroll = conn.execute("""SELECT net_salary FROM payroll WHERE employee_id=? AND month=? AND year=?""",
                           (emp_id, today.month, today.year)).fetchone()
    anns = conn.execute("""SELECT * FROM announcements WHERE is_active=1 ORDER BY ann_id DESC LIMIT 3""").fetchall()
    conn.close()
    return render_template("my_dashboard.html", emp=emp, att=att, cl=cl,
                           perf=perf, payroll=payroll, anns=anns)

# ── My Attendance (employee) ──────────────────────────────────────
@app.route("/my-attendance")
@login_required
def my_attendance():
    emp_id = session.get("employee_id")
    # Don't redirect to dashboard — that causes an infinite loop for accounts without employee records
    if not emp_id:
        return render_template("my_dashboard.html", emp=None, att=None, cl=None,
                               perf=None, payroll=None, anns=[])
    month = int(request.args.get("month", datetime.date.today().month))
    year  = int(request.args.get("year",  datetime.date.today().year))
    conn  = get_connection()
    emp   = conn.execute("SELECT name FROM employees WHERE employee_id=?", (emp_id,)).fetchone()
    records = conn.execute("""
        SELECT date, status, check_in, check_out FROM attendance
        WHERE employee_id=? AND strftime('%m',date)=? AND strftime('%Y',date)=?
        ORDER BY date
    """, (emp_id, f"{month:02d}", str(year))).fetchall()
    summary = conn.execute("""
        SELECT SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present,
               SUM(CASE WHEN status='Absent'  THEN 1 ELSE 0 END) as absent,
               SUM(CASE WHEN status='Leave'   THEN 1 ELSE 0 END) as on_leave,
               COUNT(*) as total
        FROM attendance WHERE employee_id=?
        AND strftime('%m',date)=? AND strftime('%Y',date)=?
    """, (emp_id, f"{month:02d}", str(year))).fetchone()
    conn.close()
    months = list(calendar.month_name)[1:]
    years  = list(range(2020, datetime.date.today().year + 1))
    return render_template("my_attendance.html", emp=emp, records=records,
                           summary=summary, month=month, year=year,
                           months=months, years=years,
                           month_name=calendar.month_name[month])

# ── My Payslip (employee) ──────────────────────────────────────────
@app.route("/my-payslip")
@login_required
def my_payslip():
    emp_id = session.get("employee_id")
    # Don't redirect to dashboard — that causes an infinite loop for accounts without employee records
    if not emp_id:
        return render_template("my_dashboard.html", emp=None, att=None, cl=None,
                               perf=None, payroll=None, anns=[])
    month = int(request.args.get("month", datetime.date.today().month))
    year  = int(request.args.get("year",  datetime.date.today().year))
    conn  = get_connection()
    emp   = conn.execute("""SELECT e.*, d.dept_name FROM employees e
        JOIN departments d ON e.department_id=d.department_id
        WHERE e.employee_id=?""", (emp_id,)).fetchone()
    payroll = conn.execute("""SELECT * FROM payroll WHERE employee_id=? AND month=? AND year=?""",
                           (emp_id, month, year)).fetchone()
    history = conn.execute("""SELECT month, year, net_salary, payment_status FROM payroll
        WHERE employee_id=? ORDER BY year DESC, month DESC LIMIT 12""", (emp_id,)).fetchall()
    conn.close()
    months = list(calendar.month_name)[1:]
    years  = list(range(2020, datetime.date.today().year + 1))
    history_list = [dict(h) for h in history]
    history_json = json.dumps(history_list)

    # ── Quick stats for payslip header cards ──────────────────────────────
    net_salaries = [h["net_salary"] for h in history_list if h["net_salary"]]
    avg_salary   = round(sum(net_salaries) / len(net_salaries), 2) if net_salaries else 0
    highest_pay  = max(net_salaries) if net_salaries else 0
    # Total earned this year
    conn2 = get_connection()
    ytd_row = conn2.execute("""
        SELECT COALESCE(SUM(net_salary),0) as ytd
        FROM payroll WHERE employee_id=? AND year=?
    """, (emp_id, datetime.date.today().year)).fetchone()
    conn2.close()
    total_ytd = ytd_row["ytd"] if ytd_row else 0

    # Last 6 months as pill options
    today = datetime.date.today()
    last6 = []
    for i in range(6):
        m = today.month - i
        y = today.year
        if m <= 0:
            m += 12
            y -= 1
        last6.append({"month": m, "year": y, "label": calendar.month_name[m][:3] + " " + str(y)})

    return render_template("my_payslip.html", emp=emp, payroll=payroll,
                           history=history_list, history_json=history_json,
                           month=month, year=year, months=months, years=years,
                           month_name=calendar.month_name[month],
                           avg_salary=avg_salary, highest_pay=highest_pay,
                           total_ytd=total_ytd, last6=last6)

# ── Departments ────────────────────────────────────────────────────
@app.route("/departments")
@hr_required
def departments():
    conn  = get_connection()
    depts = conn.execute("""
        SELECT d.*, COUNT(e.employee_id) as emp_count
        FROM departments d
        LEFT JOIN employees e ON e.department_id=d.department_id AND e.status='Active'
        GROUP BY d.department_id ORDER BY d.dept_name
    """).fetchall()
    conn.close()
    open_modal = request.args.get("open") == "add"
    return render_template("departments.html", departments=depts, open_modal=open_modal)

@app.route("/departments/add-form")
@hr_required
def add_department_form():
    return redirect(url_for("departments") + "?open=add")

@app.route("/departments/add", methods=["POST"])
@hr_required
def add_department():
    name    = request.form.get("dept_name","").strip()
    manager = request.form.get("manager","").strip()
    if not name:
        flash("Department name is required.", "danger")
        return redirect(url_for("departments"))
    conn = None
    try:
        conn = get_connection()
        conn.execute("INSERT INTO departments (dept_name, manager) VALUES (?,?)", (name, manager))
        conn.commit()
        flash(f"Department '{name}' added!", "success")
    except Exception as e:
        flash(f"Error: {str(e)}", "danger")
    finally:
        if conn: conn.close()
    return redirect(url_for("departments"))

@app.route("/departments/edit/<int:did>", methods=["POST"])
@hr_required
def edit_department(did):
    name    = request.form.get("dept_name","").strip()
    manager = request.form.get("manager","").strip()
    conn = None
    try:
        conn = get_connection()
        conn.execute("UPDATE departments SET dept_name=?, manager=? WHERE department_id=?",
                     (name, manager, did))
        conn.commit()
        flash("Department updated!", "success")
    except Exception as e:
        flash(f"Error: {str(e)}", "danger")
    finally:
        if conn: conn.close()
    return redirect(url_for("departments"))

@app.route("/departments/delete/<int:did>")
@hr_required
def delete_department(did):
    conn = get_connection()
    emp_count = conn.execute("SELECT COUNT(*) FROM employees WHERE department_id=? AND status='Active'",
                             (did,)).fetchone()[0]
    if emp_count > 0:
        flash(f"Cannot delete — {emp_count} active employee(s) in this department.", "danger")
        conn.close()
        return redirect(url_for("departments"))
    conn.execute("DELETE FROM departments WHERE department_id=?", (did,))
    conn.commit()
    conn.close()
    flash("Department deleted.", "info")
    return redirect(url_for("departments"))

# ── Employee Edit ──────────────────────────────────────────────────
@app.route("/employees/edit/<int:eid>", methods=["POST"])
@hr_required
def edit_employee(eid):
    name     = request.form.get("name","").strip()
    email    = request.form.get("email","").strip()
    phone    = request.form.get("phone","").strip()
    position = request.form.get("position","").strip()
    dept_id  = request.form.get("department_id")
    salary   = request.form.get("base_salary", 0)
    status   = request.form.get("status","Active")
    gender   = request.form.get("gender", "")
    conn = None
    try:
        conn = get_connection()
        conn.execute("""UPDATE employees SET name=?,email=?,phone=?,position=?,
                        department_id=?,base_salary=?,status=?,gender=? WHERE employee_id=?""",
                     (name, email, phone, position, dept_id, float(salary), status, gender, eid))
        conn.commit()
        flash(f"Employee '{name}' updated!", "success")
    except Exception as e:
        flash(f"Error: {str(e)}", "danger")
    finally:
        if conn: conn.close()
    return redirect(url_for("employees"))

# ── Announcements ──────────────────────────────────────────────────
@app.route("/announcements")
@login_required
def announcements():
    conn = get_connection()
    role = session.get("role")
    anns = conn.execute("""SELECT * FROM announcements
        WHERE is_active=1 ORDER BY ann_id DESC""").fetchall()
    all_anns = conn.execute("SELECT * FROM announcements ORDER BY ann_id DESC").fetchall() \
               if role in ("admin","hr") else []
    conn.close()
    return render_template("announcements.html", announcements=anns, all_announcements=all_anns)

@app.route("/announcements/add", methods=["POST"])
@hr_required
def add_announcement():
    title    = request.form.get("title","").strip()
    body     = request.form.get("body","").strip()
    category = request.form.get("category","General")
    priority = request.form.get("priority","Normal")
    if not title or not body:
        flash("Title and message are required.", "danger")
        return redirect(url_for("announcements"))
    conn = get_connection()
    conn.execute("""INSERT INTO announcements (title,body,category,priority,created_by,created_at,is_active)
                    VALUES (?,?,?,?,?,?,1)""",
                 (title, body, category, priority, session.get("username"),
                  str(datetime.datetime.now())))
    conn.commit()
    conn.close()
    flash("Announcement posted!", "success")
    return redirect(url_for("announcements"))

@app.route("/announcements/delete/<int:aid>")
@hr_required
def delete_announcement(aid):
    conn = get_connection()
    conn.execute("UPDATE announcements SET is_active=0 WHERE ann_id=?", (aid,))
    conn.commit()
    conn.close()
    flash("Announcement removed.", "info")
    return redirect(url_for("announcements"))

# ── Dashboard API: stat detail lists ──────────────────────────────
@app.route("/api/stat/<stat_type>")
@hr_required
def api_stat(stat_type):
    conn  = get_connection()
    today = str(datetime.date.today())
    now   = datetime.date.today()
    data  = []
    if stat_type == "total_emp":
        rows = conn.execute("""SELECT e.name, d.dept_name, e.position, e.status
            FROM employees e JOIN departments d ON e.department_id=d.department_id
            WHERE e.status='Active' ORDER BY e.name""").fetchall()
        data = [dict(r) for r in rows]
    elif stat_type == "present":
        rows = conn.execute("""SELECT e.name, d.dept_name, a.check_in, a.check_out
            FROM attendance a JOIN employees e ON a.employee_id=e.employee_id
            JOIN departments d ON e.department_id=d.department_id
            WHERE a.date=? AND a.status='Present' ORDER BY e.name""", (today,)).fetchall()
        data = [dict(r) for r in rows]
    elif stat_type == "on_leave":
        rows = conn.execute("""SELECT e.name, d.dept_name, a.status
            FROM attendance a JOIN employees e ON a.employee_id=e.employee_id
            JOIN departments d ON e.department_id=d.department_id
            WHERE a.date=? AND a.status='Leave' ORDER BY e.name""", (today,)).fetchall()
        data = [dict(r) for r in rows]
    elif stat_type == "pending_leaves":
        rows = conn.execute("""SELECT e.name, l.leave_type, l.start_date, l.end_date, l.reason
            FROM leaves l JOIN employees e ON l.employee_id=e.employee_id
            WHERE l.status='Pending' ORDER BY l.leave_id DESC""").fetchall()
        data = [dict(r) for r in rows]
    elif stat_type == "pending_payroll":
        rows = conn.execute("""SELECT e.name, d.dept_name, e.base_salary
            FROM employees e JOIN departments d ON e.department_id=d.department_id
            LEFT JOIN payroll p ON p.employee_id=e.employee_id AND p.month=? AND p.year=?
            WHERE e.status='Active' AND p.payroll_id IS NULL ORDER BY e.name""",
            (now.month, now.year)).fetchall()
        data = [dict(r) for r in rows]
    elif stat_type == "departments":
        rows = conn.execute("""SELECT d.dept_name, d.manager, COUNT(e.employee_id) as emp_count
            FROM departments d LEFT JOIN employees e ON e.department_id=d.department_id AND e.status='Active'
            GROUP BY d.department_id ORDER BY d.dept_name""").fetchall()
        data = [dict(r) for r in rows]
    elif stat_type == "approved_leave":
        rows = conn.execute("""SELECT e.name, l.leave_type, l.start_date, l.end_date
            FROM leaves l JOIN employees e ON l.employee_id=e.employee_id
            WHERE l.status='Approved' ORDER BY l.leave_id DESC LIMIT 20""").fetchall()
        data = [dict(r) for r in rows]
    elif stat_type == "absent":
        rows = conn.execute("""SELECT e.name, d.dept_name, a.date
            FROM attendance a JOIN employees e ON a.employee_id=e.employee_id
            JOIN departments d ON e.department_id=d.department_id
            WHERE a.date=? AND a.status='Absent' ORDER BY e.name""", (today,)).fetchall()
        data = [dict(r) for r in rows]
    elif stat_type == "total_ann":
        rows = conn.execute("""SELECT title, category, priority, created_by, created_at
            FROM announcements WHERE is_active=1 ORDER BY ann_id DESC""").fetchall()
        data = [dict(r) for r in rows]
    conn.close()
    return jsonify(data)


# ── Auto-migrate: ensure gender + experience columns exist ───────────────────
def ensure_extra_columns():
    conn = get_connection()
    for col, definition in [
        ("gender",           "TEXT DEFAULT ''"),
        ("experience_years", "INTEGER DEFAULT 0"),
    ]:
        try:
            conn.execute(f"ALTER TABLE employees ADD COLUMN {col} {definition}")
            conn.commit()
            print(f"✅ Added column: {col}")
        except Exception:
            pass  # Already exists — safe to ignore
    conn.close()

ensure_extra_columns()

# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    setup_database()
    app.run(debug=True, host='0.0.0.0', port=5000)


# ════════════════════════════════════════════════════════════════════
#  ADMIN — USER MANAGEMENT
# ════════════════════════════════════════════════════════════════════



# ── Admin: User Management ────────────────────────────────────────────────────
@admin_required
def admin_users():
    conn = get_connection()
    users = conn.execute("""
        SELECT u.user_id, u.username, u.email, u.role,
               COALESCE(u.status, 'Active') as status,
               u.employee_id,
               e.name  as emp_name,
               e.position as emp_position,
               d.dept_name
        FROM users u
        LEFT JOIN employees e ON u.employee_id = e.employee_id
        LEFT JOIN departments d ON e.department_id = d.department_id
        ORDER BY
          CASE u.role WHEN 'admin' THEN 1 WHEN 'hr' THEN 2 ELSE 3 END,
          u.username
    """).fetchall()

    # Employees with NO linked user account yet
    unlinked_emps = conn.execute("""
        SELECT e.employee_id, e.name, e.position, d.dept_name
        FROM employees e
        JOIN departments d ON e.department_id = d.department_id
        WHERE e.status = 'Active'
          AND e.employee_id NOT IN (
              SELECT employee_id FROM users WHERE employee_id IS NOT NULL
          )
        ORDER BY e.name
    """).fetchall()

    # All active employees (for linking)
    all_emps = conn.execute("""
        SELECT e.employee_id, e.name, e.position, d.dept_name
        FROM employees e
        JOIN departments d ON e.department_id = d.department_id
        WHERE e.status = 'Active' ORDER BY e.name
    """).fetchall()

    conn.close()
    return render_template("admin_users.html",
                           users=users,
                           unlinked_emps=unlinked_emps,
                           all_emps=all_emps)


@app.route("/admin/users/create", methods=["POST"])
@admin_required
def admin_create_user():
    username    = request.form.get("username", "").strip().lower()
    email       = request.form.get("email", "").strip().lower()
    password    = request.form.get("password", "").strip()
    role        = request.form.get("role", "employee")
    employee_id = request.form.get("employee_id") or None
    if employee_id:
        employee_id = int(employee_id)

    if not username or not password:
        flash("Username and password are required.", "danger")
        return redirect(url_for("admin_users"))

    conn = None
    try:
        conn = get_connection()
        existing = conn.execute(
            "SELECT user_id FROM users WHERE username=? OR email=?",
            (username, email)
        ).fetchone()
        if existing:
            flash(f"Username or email already exists.", "danger")
            return redirect(url_for("admin_users"))
        conn.execute(
            "INSERT INTO users (username, email, password_hash, role, employee_id, status) VALUES (?,?,?,?,?,'Active')",
            (username, email, hash_password(password), role, employee_id)
        )
        # Auto-create CL balance if linking to an employee
        if employee_id and role == "employee":
            conn.execute(
                "INSERT OR IGNORE INTO cl_balance (employee_id, year, total_cl, used_cl) VALUES (?,?,12,0)",
                (employee_id, datetime.date.today().year)
            )
        conn.commit()
        flash(f"User '{username}' created with role '{role}'. Login: {username} / {password}", "success")
    except Exception as e:
        flash(f"Error creating user: {str(e)}", "danger")
    finally:
        if conn: conn.close()
    return redirect(url_for("admin_users"))


@app.route("/admin/users/change-role/<int:uid>", methods=["POST"])
@admin_required
def admin_change_role(uid):
    new_role = request.form.get("role", "employee")
    # Prevent self-demotion if last admin
    conn = get_connection()
    if uid == session["user_id"] and new_role != "admin":
        admin_count = conn.execute(
            "SELECT COUNT(*) FROM users WHERE role='admin' AND status='Active'"
        ).fetchone()[0]
        if admin_count <= 1:
            flash("Cannot demote the only admin account.", "danger")
            conn.close()
            return redirect(url_for("admin_users"))
    conn.execute("UPDATE users SET role=? WHERE user_id=?", (new_role, uid))
    conn.commit()
    uname = conn.execute("SELECT username FROM users WHERE user_id=?", (uid,)).fetchone()["username"]
    conn.close()
    flash(f"'{uname}' is now {new_role}.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/link-employee/<int:uid>", methods=["POST"])
@admin_required
def admin_link_employee(uid):
    emp_id = request.form.get("employee_id") or None
    if emp_id:
        emp_id = int(emp_id)
    conn = get_connection()
    conn.execute("UPDATE users SET employee_id=? WHERE user_id=?", (emp_id, uid))
    conn.commit()
    uname = conn.execute("SELECT username FROM users WHERE user_id=?", (uid,)).fetchone()["username"]
    if emp_id:
        emp_name = conn.execute("SELECT name FROM employees WHERE employee_id=?", (emp_id,)).fetchone()
        conn.execute(
            "INSERT OR IGNORE INTO cl_balance (employee_id, year, total_cl, used_cl) VALUES (?,?,12,0)",
            (emp_id, datetime.date.today().year)
        )
        conn.commit()
        flash(f"'{uname}' linked to {emp_name['name'] if emp_name else emp_id}.", "success")
    else:
        flash(f"'{uname}' unlinked from employee record.", "info")
    conn.close()
    return redirect(url_for("admin_users"))


@app.route("/admin/users/reset-password/<int:uid>", methods=["POST"])
@admin_required
def admin_reset_user_password(uid):
    new_pass = request.form.get("new_password", "").strip()
    if len(new_pass) < 6:
        flash("Password must be at least 6 characters.", "danger")
        return redirect(url_for("admin_users"))
    conn = get_connection()
    conn.execute(
        "UPDATE users SET password_hash=?, reset_token=NULL, reset_expires=NULL WHERE user_id=?",
        (hash_password(new_pass), uid)
    )
    conn.commit()
    uname = conn.execute("SELECT username FROM users WHERE user_id=?", (uid,)).fetchone()["username"]
    conn.close()
    flash(f"Password for '{uname}' has been reset.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/toggle-status/<int:uid>")
@admin_required
def admin_toggle_user_status(uid):
    if uid == session["user_id"]:
        flash("You cannot deactivate your own account.", "danger")
        return redirect(url_for("admin_users"))
    conn = get_connection()
    current = conn.execute(
        "SELECT username, COALESCE(status,'Active') as status FROM users WHERE user_id=?", (uid,)
    ).fetchone()
    new_status = "Inactive" if current["status"] == "Active" else "Active"
    conn.execute("UPDATE users SET status=? WHERE user_id=?", (new_status, uid))
    conn.commit()
    conn.close()
    flash(f"'{current['username']}' is now {new_status}.", "info")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/delete/<int:uid>")
@admin_required
def admin_delete_user(uid):
    if uid == session["user_id"]:
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("admin_users"))
    conn = get_connection()
    uname = conn.execute("SELECT username FROM users WHERE user_id=?", (uid,)).fetchone()
    conn.execute("DELETE FROM users WHERE user_id=?", (uid,))
    conn.commit()
    conn.close()
    flash(f"User '{uname['username'] if uname else uid}' deleted.", "info")
    return redirect(url_for("admin_users"))


