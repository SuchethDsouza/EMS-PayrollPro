"""
export_engine.py  –  PDF payslips and Excel reports
"""
import io, calendar, datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                Paragraph, Spacer, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from database import get_connection

# ── Shared colours (used by Excel too) ───────────────────────────────────────
PURPLE = colors.HexColor("#7F77DD")
GREEN  = colors.HexColor("#1D9E75")
DARK   = colors.HexColor("#2D2B55")
LIGHT  = colors.HexColor("#F2F3F8")
RED    = colors.HexColor("#D85A30")
WHITE  = colors.white
BLACK  = colors.HexColor("#222222")


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  ENTERPRISE PDF PAYSLIP                                                     ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
def generate_payslip_pdf(employee_id, month, year):
    """Returns a BytesIO containing the redesigned enterprise PDF payslip."""
    # ── Fetch data ────────────────────────────────────────────────────────────
    conn = get_connection()
    emp = conn.execute("""
        SELECT e.*, d.dept_name FROM employees e
        JOIN departments d ON e.department_id=d.department_id
        WHERE e.employee_id=?
    """, (employee_id,)).fetchone()

    saved = conn.execute("""
        SELECT basic_pay, overtime_pay, deductions, net_salary
        FROM payroll WHERE employee_id=? AND month=? AND year=?
    """, (employee_id, month, year)).fetchone()

    att = conn.execute("""
        SELECT
            SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present,
            SUM(CASE WHEN status='Absent'  THEN 1 ELSE 0 END) as absent
        FROM attendance
        WHERE employee_id=? AND strftime('%m',date)=? AND strftime('%Y',date)=?
    """, (employee_id, f"{month:02d}", str(year))).fetchone()
    conn.close()

    if not emp or not saved:
        return None

    working_days = sum(1 for d in range(1, calendar.monthrange(year, month)[1] + 1)
                       if datetime.date(year, month, d).weekday() < 5)
    present_days = att["present"] if att else 0
    absent_days  = att["absent"]  if att else 0

    basic_pay    = saved["basic_pay"]
    overtime_pay = saved["overtime_pay"]
    deductions   = saved["deductions"]
    net_salary   = saved["net_salary"]
    gross_salary = basic_pay + overtime_pay

    def rs(amount):
        return f"Rs. {amount:,.2f}"

    # ── Colour palette ────────────────────────────────────────────────────────
    C_DARK    = colors.HexColor("#1E1B4B")
    C_ACCENT  = colors.HexColor("#4F46E5")
    C_GREEN   = colors.HexColor("#059669")
    C_RED     = colors.HexColor("#DC2626")
    C_GREY1   = colors.HexColor("#F8F9FC")
    C_GREY2   = colors.HexColor("#EEF0F6")
    C_BORDER  = colors.HexColor("#E2E5F0")
    C_TEXT1   = colors.HexColor("#111827")
    C_TEXT2   = colors.HexColor("#6B7280")
    C_WHITE   = colors.white
    C_INDIGO  = colors.HexColor("#818CF8")
    C_TEAL    = colors.HexColor("#6EE7B7")
    C_REDBG   = colors.HexColor("#FEE2E2")
    C_GREENBG = colors.HexColor("#D1FAE5")
    C_STRIP2  = colors.HexColor("#6366F1")
    C_STRIP3  = colors.HexColor("#059669")

    # ── Page setup ────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    PAGE_W, PAGE_H = A4
    LM = RM = 1.8 * cm
    W  = PAGE_W - LM - RM        # usable width in points

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=0, bottomMargin=1.5*cm,
        leftMargin=LM, rightMargin=RM
    )
    styles = getSampleStyleSheet()

    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    story = []

    # ══════════════════════════════════════════════════════════════════════════
    # 1. TOP ACCENT STRIP
    # ══════════════════════════════════════════════════════════════════════════
    strip = Table([["", "", ""]], colWidths=[W*0.5, W*0.3, W*0.2], rowHeights=[5])
    strip.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(0,0), C_ACCENT),
        ("BACKGROUND", (1,0),(1,0), C_STRIP2),
        ("BACKGROUND", (2,0),(2,0), C_STRIP3),
        ("TOPPADDING",   (0,0),(-1,-1), 0),
        ("BOTTOMPADDING",(0,0),(-1,-1), 0),
        ("LEFTPADDING",  (0,0),(-1,-1), 0),
        ("RIGHTPADDING", (0,0),(-1,-1), 0),
    ]))
    story.append(strip)

    # ══════════════════════════════════════════════════════════════════════════
    # 2. HEADER BANNER
    # ══════════════════════════════════════════════════════════════════════════
    hdr = Table([
        [
            Paragraph(
                '<font color="white" size="22"><b>PayrollPro</b></font><br/>'
                '<font color="#A5B4FC" size="9">Human Resource Management System</font>',
                ps("hL", leading=28)
            ),
            Paragraph(
                '<font color="white" size="13"><b>SALARY SLIP</b></font><br/>'
                f'<font color="#A5B4FC" size="9">{calendar.month_name[month].upper()} {year}</font><br/>'
                f'<font color="#818CF8" size="8">Ref: PP-{year}{month:02d}-{employee_id:04d}</font>',
                ps("hR", leading=18, alignment=TA_RIGHT)
            ),
        ]
    ], colWidths=[W*0.55, W*0.45])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), C_DARK),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 22),
        ("BOTTOMPADDING", (0,0),(-1,-1), 22),
        ("LEFTPADDING",   (0,0),(0,-1),  22),
        ("RIGHTPADDING",  (-1,0),(-1,-1),22),
    ]))
    story.append(hdr)
    story.append(Spacer(1, 0.5*cm))

    # ══════════════════════════════════════════════════════════════════════════
    # 3. SECTION TITLE helper
    # ══════════════════════════════════════════════════════════════════════════
    def section_title(text):
        story.append(Paragraph(
            f'<font color="#4F46E5" size="9"><b>{text}</b></font>',
            ps("sec", leading=14, spaceAfter=4)
        ))

    # ══════════════════════════════════════════════════════════════════════════
    # 4. EMPLOYEE INFO  — 4-row × 4-col: label | value || label | value
    # ══════════════════════════════════════════════════════════════════════════
    section_title("EMPLOYEE DETAILS")

    LBL = ps("lbl", fontSize=8,  textColor=C_TEXT2, leading=12)
    VAL = ps("val", fontSize=10, textColor=C_TEXT1, leading=14, fontName="Helvetica-Bold")

    def lp(text): return Paragraph(text.upper(), LBL)
    def vp(text): return Paragraph(str(text), VAL)

    # colWidths must not let padding overflow: LM/RM inside table = 10pt each
    cA, cB, cC, cD = W*0.25, W*0.25, W*0.25, W*0.25
    info_rows = [
        [lp("Employee Name"),         vp(emp["name"]),
         lp("Pay Period"),            vp(f"{calendar.month_name[month]} {year}")],
        [lp("Employee ID"),           vp(f"EMP-{emp['employee_id']:04d}"),
         lp("Working Days"),          vp(working_days)],
        [lp("Department"),            vp(emp["dept_name"]),
         lp("Present Days"),          vp(present_days)],
        [lp("Designation"),           vp(emp["position"] or "—"),
         lp("Absent Days"),           vp(absent_days)],
    ]
    info_tbl = Table(info_rows, colWidths=[cA, cB, cC, cD])
    info_tbl.setStyle(TableStyle([
        ("ROWBACKGROUNDS", (0,0),(-1,-1), [C_WHITE, C_GREY1]),
        ("TOPPADDING",    (0,0),(-1,-1), 8),
        ("BOTTOMPADDING", (0,0),(-1,-1), 8),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ("RIGHTPADDING",  (0,0),(-1,-1), 10),
        ("LINEBELOW",     (0,0),(-1,-2), 0.4, C_BORDER),
        ("LINEBEFORE",    (2,0),(2,-1),  0.8, C_BORDER),
        ("BOX",           (0,0),(-1,-1), 0.5, C_BORDER),
        # label columns muted
        ("TEXTCOLOR",     (0,0),(0,-1),  C_TEXT2),
        ("TEXTCOLOR",     (2,0),(2,-1),  C_TEXT2),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ══════════════════════════════════════════════════════════════════════════
    # 5. EARNINGS TABLE
    # ══════════════════════════════════════════════════════════════════════════
    section_title("EARNINGS")

    EH  = ps("eh",  fontSize=9,  textColor=C_WHITE, fontName="Helvetica-Bold", leading=13)
    EL  = ps("el",  fontSize=9,  textColor=C_TEXT1, leading=13)
    ELB = ps("elb", fontSize=9,  textColor=C_TEXT1, leading=13, fontName="Helvetica-Bold")
    EA  = ps("ea",  fontSize=9,  textColor=C_TEXT1, leading=13, fontName="Helvetica-Bold",
             alignment=TA_RIGHT)
    ET  = ps("et",  fontSize=10, textColor=C_ACCENT, leading=14, fontName="Helvetica-Bold",
             alignment=TA_RIGHT)

    earn = Table([
        [Paragraph("COMPONENT", EH),          Paragraph("AMOUNT", EH)],
        [Paragraph("Basic Pay", EL),           Paragraph(rs(basic_pay), EA)],
        [Paragraph("Overtime Pay", EL),        Paragraph(rs(overtime_pay), EA)],
        [Paragraph("Gross Salary", ELB),       Paragraph(rs(gross_salary), ET)],
    ], colWidths=[W*0.62, W*0.38])
    earn.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  C_ACCENT),
        ("BACKGROUND",    (0,-1),(-1,-1),C_GREY1),
        ("ROWBACKGROUNDS",(0,1),(-1,-2), [C_WHITE, C_GREY2]),
        ("TOPPADDING",    (0,0),(-1,-1), 10),
        ("BOTTOMPADDING", (0,0),(-1,-1), 10),
        ("LEFTPADDING",   (0,0),(-1,-1), 14),
        ("RIGHTPADDING",  (0,0),(-1,-1), 14),
        ("LINEBELOW",     (0,0),(-1,-2), 0.4, C_BORDER),
        ("LINEABOVE",     (0,-1),(-1,-1),0.8, C_ACCENT),
        ("BOX",           (0,0),(-1,-1), 0.5, C_BORDER),
    ]))
    story.append(earn)
    story.append(Spacer(1, 0.4*cm))

    # ══════════════════════════════════════════════════════════════════════════
    # 6. DEDUCTIONS TABLE
    # ══════════════════════════════════════════════════════════════════════════
    section_title("DEDUCTIONS")

    DH  = ps("dh",  fontSize=9,  textColor=C_WHITE, fontName="Helvetica-Bold", leading=13)
    DL  = ps("dl",  fontSize=9,  textColor=C_TEXT1, leading=13)
    DLB = ps("dlb", fontSize=9,  textColor=C_TEXT1, leading=13, fontName="Helvetica-Bold")
    DA  = ps("da",  fontSize=9,  textColor=C_RED,   leading=13, fontName="Helvetica-Bold",
             alignment=TA_RIGHT)
    DT  = ps("dt",  fontSize=10, textColor=C_RED,   leading=14, fontName="Helvetica-Bold",
             alignment=TA_RIGHT)

    ded = Table([
        [Paragraph("COMPONENT", DH),             Paragraph("AMOUNT", DH)],
        [Paragraph("Absent Day Deduction", DL),  Paragraph(f"- {rs(deductions)}", DA)],
        [Paragraph("Total Deductions", DLB),     Paragraph(f"- {rs(deductions)}", DT)],
    ], colWidths=[W*0.62, W*0.38])
    DGREY = colors.HexColor("#374151")
    ded.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  DGREY),
        ("BACKGROUND",    (0,-1),(-1,-1),C_REDBG),
        ("ROWBACKGROUNDS",(0,1),(-1,-2), [C_WHITE, C_GREY2]),
        ("TOPPADDING",    (0,0),(-1,-1), 10),
        ("BOTTOMPADDING", (0,0),(-1,-1), 10),
        ("LEFTPADDING",   (0,0),(-1,-1), 14),
        ("RIGHTPADDING",  (0,0),(-1,-1), 14),
        ("LINEBELOW",     (0,0),(-1,-2), 0.4, C_BORDER),
        ("LINEABOVE",     (0,-1),(-1,-1),0.8, C_RED),
        ("BOX",           (0,0),(-1,-1), 0.5, C_BORDER),
    ]))
    story.append(ded)
    story.append(Spacer(1, 0.5*cm))

    # ══════════════════════════════════════════════════════════════════════════
    # 7. NET SALARY HERO CARD
    # ══════════════════════════════════════════════════════════════════════════
    NL = ps("nl", fontSize=10, textColor=C_GREENBG, fontName="Helvetica-Bold", leading=14)
    NN = ps("nn", fontSize=8,  textColor=C_TEAL,    leading=12)
    NA = ps("na", fontSize=24, textColor=C_WHITE,   fontName="Helvetica-Bold",
            leading=28, alignment=TA_RIGHT)

    net = Table([
        [
            Paragraph(
                '<font color="#D1FAE5" size="10"><b>NET SALARY PAYABLE</b></font><br/>'
                f'<font color="#6EE7B7" size="8">For {calendar.month_name[month]} {year}'
                f'  |  Present: {present_days}/{working_days} days</font>',
                ps("nleft", leading=18)
            ),
            Paragraph(rs(net_salary), NA),
        ]
    ], colWidths=[W*0.52, W*0.48])
    net.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), C_GREEN),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 20),
        ("BOTTOMPADDING", (0,0),(-1,-1), 20),
        ("LEFTPADDING",   (0,0),(0,-1),  22),
        ("RIGHTPADDING",  (-1,0),(-1,-1),22),
    ]))
    story.append(net)
    story.append(Spacer(1, 0.45*cm))

    # ══════════════════════════════════════════════════════════════════════════
    # 8. SUMMARY STRIP  — Gross | Deductions | Net
    # ══════════════════════════════════════════════════════════════════════════
    cw3 = W / 3
    sum_tbl = Table([
        [
            Paragraph(
                f'<font color="#4F46E5" size="12"><b>{rs(gross_salary)}</b></font><br/>'
                '<font color="#6B7280" size="8">Gross Salary</font>',
                ps("s1", alignment=TA_CENTER, leading=18)
            ),
            Paragraph(
                f'<font color="#DC2626" size="12"><b>{rs(deductions)}</b></font><br/>'
                '<font color="#6B7280" size="8">Total Deductions</font>',
                ps("s2", alignment=TA_CENTER, leading=18)
            ),
            Paragraph(
                f'<font color="#059669" size="12"><b>{rs(net_salary)}</b></font><br/>'
                '<font color="#6B7280" size="8">Net Payable</font>',
                ps("s3", alignment=TA_CENTER, leading=18)
            ),
        ]
    ], colWidths=[cw3, cw3, cw3])
    sum_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), C_GREY1),
        ("TOPPADDING",    (0,0),(-1,-1), 14),
        ("BOTTOMPADDING", (0,0),(-1,-1), 14),
        ("LEFTPADDING",   (0,0),(-1,-1), 10),
        ("RIGHTPADDING",  (0,0),(-1,-1), 10),
        ("BOX",           (0,0),(-1,-1), 0.5, C_BORDER),
        ("LINEBEFORE",    (1,0),(2,-1),  0.5, C_BORDER),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ══════════════════════════════════════════════════════════════════════════
    # 9. FOOTER
    # ══════════════════════════════════════════════════════════════════════════
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORDER))
    story.append(Spacer(1, 0.22*cm))

    foot = Table([
        [
            Paragraph(
                "This is a computer-generated payslip and does not require a physical "
                "signature. Confidential — for employee use only.",
                ps("fl", fontSize=7.5, textColor=C_TEXT2, leading=11)
            ),
            Paragraph(
                f"Generated on {datetime.date.today().strftime('%d %B %Y')}  |  PayrollPro HRMS",
                ps("fr", fontSize=7.5, textColor=C_TEXT2, leading=11, alignment=TA_RIGHT)
            ),
        ]
    ], colWidths=[W*0.6, W*0.4])
    foot.setStyle(TableStyle([
        ("TOPPADDING",    (0,0),(-1,-1), 0),
        ("BOTTOMPADDING", (0,0),(-1,-1), 0),
        ("LEFTPADDING",   (0,0),(-1,-1), 0),
        ("RIGHTPADDING",  (0,0),(-1,-1), 0),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
    ]))
    story.append(foot)

    doc.build(story)
    buf.seek(0)
    return buf


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  EXCEL PAYROLL REPORT  (unchanged)                                          ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
def generate_payroll_excel(month, year):
    """Returns a BytesIO containing the Excel workbook."""
    conn = get_connection()
    rows = conn.execute("""
        SELECT e.employee_id, e.name, d.dept_name, e.position,
               p.basic_pay, p.overtime_pay, p.deductions, p.net_salary, p.payment_status
        FROM payroll p
        JOIN employees e ON p.employee_id=e.employee_id
        JOIN departments d ON e.department_id=d.department_id
        WHERE p.month=? AND p.year=?
        ORDER BY d.dept_name, e.name
    """, (month, year)).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll {calendar.month_abbr[month]} {year}"
    ws.sheet_view.showGridLines = False

    purple_fill = PatternFill("solid", fgColor="7F77DD")
    light_fill  = PatternFill("solid", fgColor="F2F3F8")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="E2E4EC"),
        right=Side(style="thin", color="E2E4EC"),
        top=Side(style="thin", color="E2E4EC"),
        bottom=Side(style="thin", color="E2E4EC"),
    )

    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = f"PayrollPro — Payroll Report  |  {calendar.month_name[month]} {year}"
    title.font      = Font(bold=True, size=14, color="FFFFFF")
    title.fill      = PatternFill("solid", fgColor="2D2B55")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:I2")
    gen = ws["A2"]
    gen.value = f"Generated on {datetime.date.today().strftime('%d %B %Y')} by PayrollPro"
    gen.font      = Font(size=9, color="9B97CC")
    gen.fill      = PatternFill("solid", fgColor="2D2B55")
    gen.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    headers    = ["#", "Employee Name", "Department", "Position",
                  "Basic Pay (Rs.)", "Overtime (Rs.)", "Deductions (Rs.)",
                  "Net Salary (Rs.)", "Status"]
    col_widths = [5, 22, 18, 18, 15, 14, 16, 16, 12]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = Font(bold=True, size=10, color="FFFFFF")
        cell.fill      = purple_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[3].height = 24

    for i, r in enumerate(rows, 1):
        row_num = i + 3
        fill    = white_fill if i % 2 == 0 else light_fill
        values  = [i, r["name"], r["dept_name"], r["position"],
                   r["basic_pay"], r["overtime_pay"], r["deductions"],
                   r["net_salary"], r["payment_status"]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill   = fill
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center" if col in (1,9) else "left",
                vertical="center"
            )
            if col in (5, 6, 7, 8):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row_num].height = 22

    total_row = len(rows) + 4
    ws.merge_cells(f"A{total_row}:D{total_row}")
    total_label = ws[f"A{total_row}"]
    total_label.value     = "TOTAL"
    total_label.font      = Font(bold=True, size=11, color="FFFFFF")
    total_label.fill      = PatternFill("solid", fgColor="1D9E75")
    total_label.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[total_row].height = 26

    for col, field in enumerate(["basic_pay","overtime_pay","deductions","net_salary"], 5):
        total = sum(r[field] for r in rows)
        cell  = ws.cell(row=total_row, column=col, value=total)
        cell.font          = Font(bold=True, size=11, color="FFFFFF")
        cell.fill          = PatternFill("solid", fgColor="1D9E75")
        cell.number_format = '#,##0.00'
        cell.alignment     = Alignment(horizontal="right", vertical="center")
        cell.border        = thin_border

    ws.cell(row=total_row, column=9).fill = PatternFill("solid", fgColor="1D9E75")

    # ── Attendance sheet ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Attendance Summary")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:F1")
    t2 = ws2["A1"]
    t2.value     = f"Attendance Summary — {calendar.month_name[month]} {year}"
    t2.font      = Font(bold=True, size=13, color="FFFFFF")
    t2.fill      = PatternFill("solid", fgColor="2D2B55")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    att_headers = ["Employee", "Department", "Present", "Absent", "Leave", "Attendance %"]
    att_widths  = [22, 18, 12, 12, 12, 15]
    for col, (h, w) in enumerate(zip(att_headers, att_widths), 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font      = Font(bold=True, size=10, color="FFFFFF")
        cell.fill      = purple_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
        ws2.column_dimensions[cell.column_letter].width = w
    ws2.row_dimensions[2].height = 22

    conn = get_connection()
    att_rows = conn.execute("""
        SELECT e.name, d.dept_name,
               SUM(CASE WHEN a.status='Present' THEN 1 ELSE 0 END) as present,
               SUM(CASE WHEN a.status='Absent'  THEN 1 ELSE 0 END) as absent,
               SUM(CASE WHEN a.status='Leave'   THEN 1 ELSE 0 END) as on_leave
        FROM attendance a
        JOIN employees e ON a.employee_id=e.employee_id
        JOIN departments d ON e.department_id=d.department_id
        WHERE strftime('%m',a.date)=? AND strftime('%Y',a.date)=?
        GROUP BY e.employee_id ORDER BY e.name
    """, (f"{month:02d}", str(year))).fetchall()
    conn.close()

    for i, r in enumerate(att_rows, 1):
        row_num   = i + 2
        fill      = white_fill if i % 2 == 0 else light_fill
        total_days = r["present"] + r["absent"] + r["on_leave"]
        pct        = round(r["present"] / total_days * 100, 1) if total_days else 0
        for col, val in enumerate([r["name"], r["dept_name"], r["present"],
                                   r["absent"], r["on_leave"], f"{pct}%"], 1):
            cell = ws2.cell(row=row_num, column=col, value=val)
            cell.fill   = fill
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center" if col > 2 else "left",
                vertical="center"
            )
            if col == 6 and pct < 75:
                cell.font = Font(color="D85A30", bold=True)
            elif col == 6:
                cell.font = Font(color="1D9E75", bold=True)
        ws2.row_dimensions[row_num].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
